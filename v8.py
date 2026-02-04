#This is v9
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
from datetime import datetime
import base64
import os 

# PDF imports
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Excel imports
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import comments
    try:
        from openpyxl.worksheet.data_validation import DataValidation
        HAS_DATA_VALIDATION = True
    except ImportError:
        HAS_DATA_VALIDATION = False
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    HAS_DATA_VALIDATION = False

# PAGE CONFIG
if 'page_config_set' not in st.session_state:
    st.set_page_config(
        page_title="Proaltus - Análisis de Portafolio",
        page_icon="💎",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    st.session_state.page_config_set = True

# AUTHENTICATION SYSTEM
def check_authentication():
    return st.session_state.get('authenticated', False)

def show_login():
    st.markdown("""
    <style>
        .login-container {
            max-width: 400px;
            margin: 0 auto;
            padding: 3rem;
            background: white;
            border-radius: 20px;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
            border: 1px solid #E5E7EB;
            margin-top: 10%;
        }
        .login-header {
            text-align: center;
            margin-bottom: 2rem;
        }
        .login-title {
            font-size: 2.5rem;
            font-weight: 700;
            color: #1E3A8A;
            margin-bottom: 0.5rem;
            font-family: 'Inter', sans-serif;
        }
        .login-subtitle {
            color: #6B7280;
            font-size: 1rem;
            margin-bottom: 2rem;
        }
    </style>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("""
        <div class="login-container">
            <div class="login-header">
                <h1 class="login-title">PROALTUS</h1>
                <p class="login-subtitle">Sistema de Análisis de Portafolio</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("login_form"):
            st.markdown("### Acceso Seguro")
            username = st.text_input("Usuario", placeholder="Ingrese su usuario")
            password = st.text_input("Contraseña", type="password", placeholder="Ingrese su contraseña")
            submit_button = st.form_submit_button("INGRESAR", type="primary", use_container_width=True)
            
            if submit_button:
                if username == "Proaltus" and password == "Proaltus2025":
                    st.session_state.authenticated = True
                    st.success("Acceso autorizado! Bienvenido al sistema.")
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos")

if not check_authentication():
    show_login()
    st.stop()

# CORPORATE CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600&display=swap');
    
    :root {
        --primary-blue: #1E3A8A;
        --secondary-blue: #3B82F6;
        --light-blue: #DBEAFE;
        --accent-blue: #1D4ED8;
        --dark-gray: #1F2937;
        --medium-gray: #6B7280;
        --light-gray: #F3F4F6;
        --border-gray: #E5E7EB;
        --white: #FFFFFF;
        --success-green: #059669;
        --warning-orange: #D97706;
        --error-red: #DC2626;
        --gradient-primary: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
        --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
        --shadow-2xl: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        --radius-lg: 12px;
        --radius-xl: 16px;
    }
    
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        color: var(--dark-gray);
        background: var(--light-gray);
    }
    
    .main .block-container {
        background: var(--light-gray);
        padding-top: 2rem;
        padding-bottom: 2rem;
        max-width: 1400px;
    }
    
    .corporate-header {
        background: var(--gradient-primary);
        padding: 3rem 2rem 2rem 2rem;
        border-radius: var(--radius-xl);
        margin-bottom: 2rem;
        color: var(--white);
        position: relative;
        overflow: hidden;
        box-shadow: var(--shadow-2xl);
    }
    
    .header-title {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.1);
        letter-spacing: -0.025em;
    }
    
    .header-subtitle {
        font-size: 1.25rem;
        font-weight: 400;
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        letter-spacing: 0.025em;
    }
    
    .section-container {
        background: var(--white);
        border-radius: var(--radius-lg);
        padding: 2rem;
        margin: 2rem 0;
        box-shadow: var(--shadow-md);
        border: 1px solid var(--border-gray);
    }
    
    .section-title {
        font-size: 1.5rem;
        font-weight: 600;
        color: var(--primary-blue);
        margin: 0;
        letter-spacing: -0.025em;
    }
    
    .kpi-card {
        background: var(--white);
        border-radius: var(--radius-lg);
        padding: 2rem;
        box-shadow: var(--shadow-lg);
        border: 1px solid var(--border-gray);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    
    .kpi-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: var(--gradient-primary);
    }
    
    .kpi-card:hover {
        transform: translateY(-4px);
        box-shadow: var(--shadow-2xl);
    }
    
    .kpi-title {
        font-size: 0.875rem;
        font-weight: 600;
        color: var(--medium-gray);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin-bottom: 1rem;
    }
    
    .kpi-value {
        font-size: 2.5rem;
        font-weight: 700;
        color: var(--primary-blue);
        margin-bottom: 0.5rem;
        line-height: 1.1;
        font-family: 'JetBrains Mono', monospace;
    }
    
    .kpi-meta {
        font-size: 0.75rem;
        color: var(--medium-gray);
        margin-top: 0.5rem;
        font-family: 'JetBrains Mono', monospace;
    }
    
    .status-indicator {
        display: inline-flex;
        align-items: center;
        padding: 0.5rem 1rem;
        border-radius: 8px;
        font-size: 0.875rem;
        font-weight: 500;
        gap: 0.5rem;
    }
    
    .status-success {
        background: rgba(5, 150, 105, 0.1);
        color: var(--success-green);
        border: 1px solid rgba(5, 150, 105, 0.2);
    }
    
    .status-warning {
        background: rgba(217, 119, 6, 0.1);
        color: var(--warning-orange);
        border: 1px solid rgba(217, 119, 6, 0.2);
    }
    
    .status-dot {
        width: 8px;
        height: 8px;
        border-radius: 50%;
        background: currentColor;
    }
    
    .stButton > button {
        background: var(--gradient-primary);
        color: var(--white);
        border: none;
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 0.875rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        box-shadow: var(--shadow-md);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: var(--shadow-lg);
        background: var(--primary-blue);
    }
    
    .dataframe thead th {
        background: var(--primary-blue) !important;
        color: var(--white) !important;
        font-weight: 600 !important;
        padding: 1rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.05em !important;
    }
    
    .dataframe tbody td {
        padding: 1rem !important;
        border-bottom: 1px solid var(--border-gray) !important;
        font-family: 'JetBrains Mono', monospace !important;
    }
    
    .dataframe tbody tr:nth-child(even) {
        background: var(--light-gray) !important;
    }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# UTILITY FUNCTIONS
def safe_float(value, default=0):
    """Safely convert value to float with better error handling"""
    try:
        if pd.isna(value) or value is None or value == '':
            return default
        if isinstance(value, str):
            value = value.replace(',', '').replace('$', '').replace('%', '').strip()
            if value == '' or value.lower() in ['nan', 'none', 'null']:
                return default
        return float(value)
    except (ValueError, TypeError):
        return default

def find_exact_column(df, exact_names):
    """Find column by exact name matches"""
    for exact_name in exact_names:
        for col in df.columns:
            if str(col).strip() == exact_name:
                return col
    return None

def validate_dataframe(df, required_columns):
    """Validate that dataframe has required columns"""
    if df is None or df.empty:
        return False, "DataFrame is empty or None"
    
    missing_cols = []
    for req_col in required_columns:
        if find_exact_column(df, [req_col]) is None:
            missing_cols.append(req_col)
    
    if missing_cols:
        return False, f"Missing columns: {missing_cols}"
    
    return True, "OK"

# DATA PROCESSING FUNCTIONS
def process_with_pandas(uploaded_file):
    """Process Excel file with better error handling and exact column matching"""
    try:
        processed_data = {}
        engines_to_try = ['openpyxl', 'xlrd', None]
        excel_data = None
        
        for engine in engines_to_try:
            try:
                if engine == 'openpyxl' and not OPENPYXL_AVAILABLE:
                    continue
                excel_data = pd.read_excel(uploaded_file, sheet_name=None, engine=engine)
                break
            except Exception:
                continue
        
        if excel_data is None:
            raise Exception("Could not read Excel file with any available engine")
        
        sheets_config = {
            'Empresas': {
                'key': 'empresas',
                'valor_col': 'Valor Patrimonial (COP)',
                'nombre_col': 'Nombre'
            },
            'Inversiones No Productivas': {
                'key': 'inversiones_no_productivas',
                'valor_col': 'Valor (COP)',
                'nombre_col': 'Nombre del Activo'
            },
            'Inversiones Productivas': {
                'key': 'inversiones_productivas',
                'valor_col': 'Valor (COP)',
                'nombre_col': 'Nombre del Proyecto'
            },
            'Inversiones Financieras': {
                'key': 'inversiones_financieras',
                'valor_col': 'Valor (COP)',
                'nombre_col': 'Nombre del Activo'
            },
            'Datos adicionales': {
                'key': 'datos_adicionales',
                'valor_col': 'Valor (COP)',
                'categoria_col': 'Categoría',
                'subcategoria_col': 'Subcategoria ',
                'tipo_col': 'Tipo de Relación'
            }
        }
        
        for sheet_name, config in sheets_config.items():
            if sheet_name in excel_data:
                df = excel_data[sheet_name].copy()
                
                df = df.dropna(how='all')
                df = df[~df.iloc[:, 0].astype(str).str.upper().str.contains('TOTAL', na=False)]
                
                if not df.empty:
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    valor_col = find_exact_column(df, [config['valor_col']])
                    if valor_col:
                        df[valor_col] = pd.to_numeric(df[valor_col], errors='coerce').fillna(0)
                    
                    if config['key'] == 'datos_adicionales':
                        required_cols = ['Categoría', 'Subcategoria ', 'Valor (COP)', 'Tipo de Relación']
                        is_valid, msg = validate_dataframe(df, required_cols)
                        if not is_valid:
                            st.warning(f"Datos adicionales: {msg}")
                            continue
                    
                    processed_data[config['key']] = df
        
        return processed_data
        
    except Exception as e:
        st.error(f"Error processing with pandas: {str(e)}")
        return None

def process_uploaded_template(uploaded_file):
    """Main template processing function with improved error handling"""
    try:
        if not OPENPYXL_AVAILABLE:
            return process_with_pandas(uploaded_file)
        
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        processed_data = {}
        
        sheets_config = {
            'Empresas': 'empresas',
            'Inversiones No Productivas': 'inversiones_no_productivas',
            'Inversiones Productivas': 'inversiones_productivas', 
            'Inversiones Financieras': 'inversiones_financieras',
            'Datos adicionales': 'datos_adicionales'
        }
        
        for sheet_name, data_key in sheets_config.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                    else:
                        break
                
                data_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]).upper() != 'TOTAL':
                        row_data = row[:len(headers)]
                        data_rows.append(row_data)
                
                if data_rows and headers:
                    df = pd.DataFrame(data_rows, columns=headers)
                    df = df.dropna(how='all')
                    
                    # Convert numeric columns according to exact configuration
                    numeric_columns = {
                        'empresas': ['Valor Patrimonial (COP)', 'Porcentaje'],
                        'inversiones_no_productivas': ['Valor (COP)', 'Total_management_fee', 'Impuestos'],
                        'inversiones_productivas': ['Valor (COP)', 'Ingreso Mesual'],
                        'inversiones_financieras': ['Valor (COP)', 'Ingreso mensual '],
                        'datos_adicionales': ['Valor (COP)']
                    }
                    
                    if data_key in numeric_columns:
                        for col_name in numeric_columns[data_key]:
                            col = find_exact_column(df, [col_name])
                            if col:
                                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    
                    processed_data[data_key] = df
        
        return processed_data
        
    except Exception as e:
        st.error(f"Error processing Excel file: {str(e)}")
        return None

# MEKKO CHART FUNCTIONS
def create_proper_mekko_chart(categories, values, title, height=400, colors=None):
    """Creates a proper Mekko chart with correct proportional rectangles"""
    total = sum(values)
    if total == 0:
        st.warning(f"No data to display in {title}")
        return
    
    if colors is None:
        colors = ['#1E3A8A', '#3B82F6', '#60A5FA', '#10B981', '#34D399', '#F59E0B', '#FCD34D', '#EF4444', '#F87171', '#8B5CF6', '#A78BFA', '#EC4899'] * 3
    
    fig = go.Figure()
    
    valid_data = [(cat, val, colors[i % len(colors)]) for i, (cat, val) in enumerate(zip(categories, values)) if val > 0]
    
    if not valid_data:
        st.warning("No valid data for Mekko chart")
        return
    
    num_items = len(valid_data)
    if num_items <= 6:
        rows = [valid_data]
    elif num_items <= 12:
        mid = num_items // 2
        rows = [valid_data[:mid], valid_data[mid:]]
    else:
        third = num_items // 3
        rows = [valid_data[:third], valid_data[third:2*third], valid_data[2*third:]]
    
    row_totals = [sum(val for _, val, _ in row) for row in rows]
    total_all = sum(row_totals)
    row_heights = [(rt / total_all) * 100 for rt in row_totals]
    
    y_current = 0
    
    for row_idx, (row, row_height) in enumerate(zip(rows, row_heights)):
        if not row:
            continue
            
        row_total = sum(val for _, val, _ in row)
        x_current = 0
        
        for cat, val, color in row:
            if val > 0:
                width = (val / row_total) * 100
                
                fig.add_shape(
                    type="rect",
                    x0=x_current,
                    y0=y_current,
                    x1=x_current + width,
                    y1=y_current + row_height,
                    fillcolor=color,
                    line=dict(color="white", width=2),
                    opacity=0.9
                )
                
                if width > 8 and row_height > 10:
                    font_size = max(8, min(12, int(width / 10)))
                    fig.add_annotation(
                        x=x_current + width/2,
                        y=y_current + row_height/2,
                        text=f"<b>{cat}</b><br>${val:,.0f}<br>({val/total*100:.1f}%)",
                        showarrow=False,
                        font=dict(color="white", size=font_size, family="Inter"),
                        align="center"
                    )
                
                x_current += width
        
        y_current += row_height
    
    fig.update_layout(
        title=dict(
            text=title,
            font=dict(size=18, color='#1F2937', family="Inter"),
            x=0.5,
            xanchor='center'
        ),
        height=height,
        paper_bgcolor='white',
        plot_bgcolor='white',
        xaxis=dict(
            showgrid=False,
            showticklabels=False,
            range=[0, 100],
            fixedrange=True
        ),
        yaxis=dict(
            showgrid=False,
            showticklabels=False,
            range=[0, 100],
            fixedrange=True
        ),
        font=dict(family="Inter"),
        margin=dict(l=0, r=0, t=60, b=10),
        showlegend=False
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_expenses_mekko_chart(processed_data):
    """Creates Mekko chart with ONLY expenses from Datos adicionales"""
    
    if 'datos_adicionales' not in processed_data:
        st.warning("No additional data available for detailed expenses")
        return
    
    df_datos = processed_data['datos_adicionales']
    
    categoria_col = find_exact_column(df_datos, ['Categoría'])
    subcategoria_col = find_exact_column(df_datos, ['Subcategoria '])
    valor_col = find_exact_column(df_datos, ['Valor (COP)'])
    tipo_col = find_exact_column(df_datos, ['Tipo de Relación'])
    
    if not all([categoria_col, valor_col, tipo_col]):
        available_cols = list(df_datos.columns)
        st.warning(f"Required columns not found. Available columns: {available_cols}")
        return
    
    df_work = df_datos.copy()
    df_work[valor_col] = pd.to_numeric(df_work[valor_col], errors='coerce').fillna(0)
    
    # FILTER ONLY EXPENSES
    df_work = df_work[df_work[tipo_col] == 'Egreso']
    df_work = df_work[df_work[valor_col] > 0]
    
    if df_work.empty:
        st.warning("No valid expense data found")
        return
    
    expense_data = {}
    
    for _, row in df_work.iterrows():
        # Use subcategory if available, otherwise use category
        if subcategoria_col and pd.notna(row[subcategoria_col]) and str(row[subcategoria_col]).strip() != '':
            categoria = str(row[subcategoria_col]).strip()
        else:
            categoria = str(row[categoria_col]).strip()
        
        valor = safe_float(row[valor_col])
        
        if valor > 0 and categoria and categoria.lower() not in ['nan', '', 'none']:
            if categoria in expense_data:
                expense_data[categoria] += valor
            else:
                expense_data[categoria] = valor
    
    if not expense_data:
        st.warning("No valid expense categories found")
        return
    
    sorted_expenses = sorted(expense_data.items(), key=lambda x: x[1], reverse=True)
    categories = [item[0] for item in sorted_expenses]
    values = [item[1] for item in sorted_expenses]
    
    create_proper_mekko_chart(categories, values, "Distribución de Gastos - Mekko Chart", height=500)

def create_patrimony_mekko_chart(kpis):
    """Creates Mekko chart for patrimony distribution"""
    
    categories = [
        'Empresas',
        'Inv. Productivas', 
        'Inv. No Productivas',
        'Inv. Financieras'
    ]
    
    values = [
        safe_float(kpis.get('total_companies', 0)),
        safe_float(kpis.get('total_productive', 0)),
        safe_float(kpis.get('total_non_productive', 0)),
        safe_float(kpis.get('total_financial', 0))
    ]
    
    colors = ['#1E3A8A', '#10B981', '#F59E0B', '#8B5CF6']
    
    create_proper_mekko_chart(categories, values, "Distribución del Patrimonio - Mekko Chart", height=400, colors=colors)

# CASH FLOW ANALYSIS - CORRECTED ACCORDING TO MANUAL FORMULAS
def generate_cash_flow_analysis(data):
    """Generate comprehensive cash flow analysis following exact manual formulas - CORREGIDO"""
    try:
        flow_analysis = {}
        
        # STEP 1: CALCULATE INGRESOS FOLLOWING EQUATIONS 7-8
        ingreso_salarial = 0  # Equation 7: Sum of salaries and wages
        ingresos_pasivos = 0  # Equation 8: Income from financial and productive investments
        
        # 1.1 Ingreso Salarial from Datos adicionales
        if 'datos_adicionales' in data:
            df_datos = data['datos_adicionales']
            
            categoria_col = find_exact_column(df_datos, ['Categoría'])
            subcategoria_col = find_exact_column(df_datos, ['Subcategoria '])
            valor_col = find_exact_column(df_datos, ['Valor (COP)'])
            tipo_col = find_exact_column(df_datos, ['Tipo de Relación'])
            
            if all([categoria_col, valor_col, tipo_col]):
                # Filter only income
                ingresos_data = df_datos[df_datos[tipo_col] == 'Ingreso']
                
                for _, row in ingresos_data.iterrows():
                    valor = safe_float(row[valor_col])
                    # All income from "Datos adicionales" is considered salary income
                    ingreso_salarial += valor
        
        # 1.2 Ingresos Pasivos from investments (Equation 8) - ACTUALIZADO v2
        # From financial investments
        if 'inversiones_financieras' in data:
            df_fin = data['inversiones_financieras']
            ingreso_col = find_exact_column(df_fin, ['Rendimiento Mensual'])
            if ingreso_col:
                for _, row in df_fin.iterrows():
                    ingreso_value = safe_float(row[ingreso_col])
                    ingresos_pasivos += ingreso_value

        # From productive investments  
        if 'inversiones_productivas' in data:
            df_prod = data['inversiones_productivas']
            ingreso_col = find_exact_column(df_prod, ['Rendimiento Mensual'])
            if ingreso_col:
                for _, row in df_prod.iterrows():
                    ingreso_value = safe_float(row[ingreso_col])
                    ingresos_pasivos += ingreso_value
        
        # Total Income (Equation 3)
        total_ingresos = ingreso_salarial + ingresos_pasivos
        
        # STEP 2: CALCULATE EGRESOS FOLLOWING EQUATION 4
        # Etotal = GP1 + GP2 + INV + IMP
        
        # Initialize expense categories
        gesenciales = 0
        goperativos = 0
        gvarios = 0
        gviajes = 0
        glujo = 0
        pension_voluntaria = 0
        proyecto_inmobiliarios = 0
        provision_impuestos = 0
        
        # 2.1 Calculate expense categories from Datos adicionales
        if 'datos_adicionales' in data:
            df_datos = data['datos_adicionales']
            
            if all([categoria_col, valor_col, tipo_col]):
                egresos_data = df_datos[df_datos[tipo_col] == 'Egreso']
                
                for _, row in egresos_data.iterrows():
                    categoria = str(row[categoria_col]).strip()
                    valor = safe_float(row[valor_col])
                    
                    # Classify according to manual methodology
                    if categoria == 'Gastos Esenciales':
                        gesenciales += valor
                    elif categoria == 'Gastos Operativos':
                        goperativos += valor
                    elif categoria == 'Gastos Varios':
                        gvarios += valor
                    elif categoria == 'Viajes':
                        gviajes += valor
                    elif categoria == 'Lujo':
                        glujo += valor
                    elif categoria == 'Inversiones':
                        # CORREGIDO: Desglosar por subcategoría con mejor lógica
                        if subcategoria_col and pd.notna(row[subcategoria_col]):
                            subcategoria = str(row[subcategoria_col]).lower()
                            if any(keyword in subcategoria for keyword in ['pensión', 'pension', 'voluntaria']):
                                pension_voluntaria += valor
                            elif any(keyword in subcategoria for keyword in ['inmobiliario', 'proyecto', 'inmobiliarios']):
                                proyecto_inmobiliarios += valor
                        else:
                            # Si no hay subcategoría específica, asignar a proyectos inmobiliarios por defecto
                            proyecto_inmobiliarios += valor
                    elif categoria == 'Impuestos':
                        provision_impuestos += valor
        
       # 2.2 Calculate maintenance costs (Equation 9 and 12) - CORREGIDO v3
        cmantenimiento_mensual = 0
        impuestos_inversiones_mensual = 0

        if 'inversiones_no_productivas' in data:
            df_no_prod = data['inversiones_no_productivas']
            
            # Usar la columna 'Costo mantenimiento' (valores ANUALES, dividir entre 12)
            costo_mant_col = find_exact_column(df_no_prod, [
                'Costo mantenimiento',
                'Costo mantenimiento ',
                ' Costo mantenimiento'
            ])
            
            if costo_mant_col:
                for _, row in df_no_prod.iterrows():
                    costo_anual = safe_float(row[costo_mant_col])
                    # Convertir de anual a mensual
                    costo_mensual = costo_anual / 12
                    cmantenimiento_mensual += costo_mensual
            
            # Annual taxes converted to monthly (Equation 13)
            tax_col = find_exact_column(df_no_prod, ['Impuestos'])
            if tax_col:
                impuestos_anuales = safe_float(df_no_prod[tax_col].sum())
                impuestos_inversiones_mensual = impuestos_anuales / 12
        
        # STEP 3: CALCULATE EXPENSE CATEGORIES ACCORDING TO EQUATIONS 9-10
        # GP1 = Gesenciales + Goperativos + Cmantenimiento/12 (Equation 9)
        total_gastos_p1 = gesenciales + goperativos + cmantenimiento_mensual
        
        # GP2 = Gvarios + Gviajes + Glujo (Equation 10)  
        total_gastos_p2 = gvarios + gviajes + glujo
        
        # INV = Inversiones mensuales (pension + proyectos inmobiliarios)
        total_inversiones = pension_voluntaria + proyecto_inmobiliarios
        
        # IMP = Impuestos (Equation 13)
        total_impuestos = impuestos_inversiones_mensual + provision_impuestos
        
        # STEP 4: CALCULATE TOTALS (Equations 2-4)
        # Etotal = GP1 + GP2 + INV + IMP (Equation 4)
        total_egresos = total_gastos_p1 + total_gastos_p2 + total_inversiones + total_impuestos
        
        # FCN = Itotal - Etotal (Equation 2)
        resultado_neto = total_ingresos - total_egresos
        
        # Build complete analysis structure
        flow_analysis = {
            'ingresos': {
                'ingreso_salarial': ingreso_salarial,
                'ingresos_pasivos': ingresos_pasivos,
                'total': total_ingresos
            },
            'gastos_p1': {
                'gastos_esenciales': gesenciales,
                'gastos_operativos': goperativos,
                'relacionado_inversiones': cmantenimiento_mensual,
                'total': total_gastos_p1
            },
            'gastos_p2': {
                'gastos_varios': gvarios,
                'viajes': gviajes,
                'lujo': glujo,
                'total': total_gastos_p2
            },
            'inversiones': {
                'pension_voluntaria': pension_voluntaria,
                'proyecto_inmobiliarios': proyecto_inmobiliarios,
                'total': total_inversiones
            },
            'impuestos': {
                'impuestos_inversiones': impuestos_inversiones_mensual,
                'provision_impuestos': provision_impuestos,
                'total': total_impuestos
            },
            'resumen': {
                'total_egresos': total_egresos,
                'resultado_neto': resultado_neto,
                'porcentajes': {
                    'gastos_p1': (total_gastos_p1 / total_ingresos * 100) if total_ingresos > 0 else 0,
                    'gastos_p2': (total_gastos_p2 / total_ingresos * 100) if total_ingresos > 0 else 0,
                    'inversiones': (total_inversiones / total_ingresos * 100) if total_ingresos > 0 else 0,
                    'impuestos': (total_impuestos / total_ingresos * 100) if total_ingresos > 0 else 0,
                    'resultado_neto': (resultado_neto / total_ingresos * 100) if total_ingresos > 0 else 0
                }
            }
        }
        
        return flow_analysis
        
    except Exception as e:
        st.error(f"Error in cash flow analysis: {str(e)}")
        return None

def calculate_patrimony_kpis(data):
    """Calculate comprehensive KPIs following exact manual formulas"""
    try:
        kpis = {
            'total_patrimony': 0.0,
            'total_companies': 0.0,
            'total_non_productive': 0.0,
            'total_productive': 0.0,
            'total_financial': 0.0,
            'total_income': 0.0,
            'total_expenses': 0.0,
            'net_flow': 0.0,
            'savings_rate': 0.0,
            'asset_count': 0,
        }
        
        # EQUATION 1: Ptotal = Pempresas + Pno_productivas + Pproductivas + Pfinancieras
        
        # Calculate patrimony by categories using exact names
        if 'empresas' in data:
            df = data['empresas']
            valor_col = find_exact_column(df, ['Valor Patrimonial (COP)'])
            if valor_col:
                kpis['total_companies'] = safe_float(df[valor_col].sum())
        
        if 'inversiones_no_productivas' in data:
            df = data['inversiones_no_productivas']
            valor_col = find_exact_column(df, ['Valor (COP)'])
            if valor_col:
                kpis['total_non_productive'] = safe_float(df[valor_col].sum())
        
        if 'inversiones_productivas' in data:
            df = data['inversiones_productivas']
            valor_col = find_exact_column(df, ['Valor (COP)'])
            if valor_col:
                kpis['total_productive'] = safe_float(df[valor_col].sum())
        
        if 'inversiones_financieras' in data:
            df = data['inversiones_financieras']
            valor_col = find_exact_column(df, ['Valor (COP)'])
            if valor_col:
                kpis['total_financial'] = safe_float(df[valor_col].sum())
        
        # Total patrimony (Equation 1)
        kpis['total_patrimony'] = (kpis['total_companies'] + kpis['total_non_productive'] + 
                                  kpis['total_productive'] + kpis['total_financial'])
        
        # EQUATION 6: Count assets
        for category in ['empresas', 'inversiones_no_productivas', 'inversiones_productivas', 'inversiones_financieras']:
            if category in data and not data[category].empty:
                df = data[category]
                valid_rows = df[~df.iloc[:, 0].astype(str).str.upper().str.contains('TOTAL', na=False)]
                kpis['asset_count'] += len(valid_rows)
        
        # Use corrected cash flow analysis
        flow_analysis = generate_cash_flow_analysis(data)
        if flow_analysis:
            kpis['total_income'] = flow_analysis['ingresos']['total']
            kpis['total_expenses'] = flow_analysis['resumen']['total_egresos']
            kpis['net_flow'] = flow_analysis['resumen']['resultado_neto']
            # EQUATION 5: TA = (FCN / Itotal) × 100
            kpis['savings_rate'] = flow_analysis['resumen']['porcentajes']['resultado_neto']
        
        return kpis
        
    except Exception as e:
        st.error(f"Error calculating KPIs: {str(e)}")
        return None

# CHART FUNCTIONS
def create_cash_flow_graphic(flow_analysis):
    """Create cash flow graphic following manual specifications - Gráfica 1"""
    ingresos = flow_analysis['ingresos']
    fig = go.Figure()
    
    # Main bar (Dark Blue): Total Income - represents 100% of financial capacity
    fig.add_trace(go.Bar(
        y=['Ingreso'],
        x=[ingresos['total']],
        orientation='h',
        marker_color='#1E3A8A',
        text=[f"${ingresos['total']:,.0f}"],
        textposition='inside',
        textfont=dict(color='white', size=16, family="Inter", weight="bold"),
        name='Ingreso Total',
        width=0.6,
        hovertemplate='<b>%{y}</b><br>Valor: $%{x:,.0f}<extra></extra>'
    ))
    
    # Secondary bar (Gray): Salary Income - shows work dependency
    fig.add_trace(go.Bar(
        y=['Ingreso Salarial'],
        x=[ingresos['ingreso_salarial']],
        orientation='h',
        marker_color='#F3F4F6',
        text=[f"${ingresos['ingreso_salarial']:,.0f}"],
        textposition='inside',
        textfont=dict(color='#1F2937', size=14, family="Inter", weight="bold"),
        name='Ingreso Salarial',
        width=0.4,
        hovertemplate='<b>%{y}</b><br>Valor: $%{x:,.0f}<extra></extra>'
    ))
    
    # Tertiary bar (Medium Blue): Passive Income - indicates financial independence level
    fig.add_trace(go.Bar(
        y=['Ingresos Pasivos'],
        x=[ingresos['ingresos_pasivos']],
        orientation='h',
        marker_color='#60A5FA',
        text=[f"${ingresos['ingresos_pasivos']:,.0f}"],
        textposition='inside',
        textfont=dict(color='white', size=14, family="Inter", weight="bold"),
        name='Ingresos Pasivos',
        width=0.4,
        hovertemplate='<b>%{y}</b><br>Valor: $%{x:,.0f}<extra></extra>'
    ))
    
    fig.update_layout(
        title="",
        height=350,
        paper_bgcolor='white',
        plot_bgcolor='white',
        showlegend=False,
        margin=dict(l=140, r=50, t=30, b=50),
        xaxis=dict(
            showgrid=False, 
            showticklabels=False, 
            zeroline=False,
            range=[0, ingresos['total'] * 1.1]
        ),
        yaxis=dict(
            showgrid=False, 
            tickfont=dict(size=14, color='#1F2937', family="Inter"),
            categoryorder='array',
            categoryarray=['Ingresos Pasivos', 'Ingreso Salarial', 'Ingreso']
        ),
        font=dict(family="Inter", color='#1F2937'),
        barmode='group'
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_maintenance_costs_graphic(processed_data):
    """Create maintenance costs graphic - Gráfica 3 - CORREGIDO"""
    if 'inversiones_no_productivas' not in processed_data:
        st.warning("No non-productive investment data")
        return
    
    df_no_prod = processed_data['inversiones_no_productivas']
    
    name_col = find_exact_column(df_no_prod, ['Nombre del Activo'])
    
    # Usar la nueva columna 'Costo mantenimiento' (ya es mensual)
    costo_mant_col = find_exact_column(df_no_prod, ['Costo mantenimiento'])

    if not costo_mant_col:
        st.warning(f"'Costo mantenimiento' column not found. Available: {list(df_no_prod.columns)}")
        return

    nombres = []
    costos_mensuales = []

    df_valid = df_no_prod.copy()
    df_valid[costo_mant_col] = pd.to_numeric(df_valid[costo_mant_col], errors='coerce').fillna(0)

    for _, row in df_valid.iterrows():
        costo_anual = safe_float(row[costo_mant_col])
        # Convertir de anual a mensual
        costo_mensual = costo_anual / 12
        
        if costo_mensual > 0:
            nombres.append(str(row[name_col]))
            costos_mensuales.append(costo_mensual)
        
        if not costos_mensuales:
            st.warning("No valid maintenance cost data")
            return
    
    max_costo = max(costos_mensuales) if costos_mensuales else 0
    colors = []
    
    for i, costo in enumerate(costos_mensuales):
        if costo == max_costo:
            colors.append('#1E3A8A')
        elif costo > max_costo * 0.5:
            colors.append('#3B82F6')
        elif costo > max_costo * 0.2:
            colors.append('#60A5FA')
        else:
            colors.append('#9CA3AF')
    
    fig = go.Figure(go.Bar(
        x=nombres,
        y=costos_mensuales,
        marker_color=colors,
        text=[f"${c:,.0f}" if c > 0 else "" for c in costos_mensuales],
        textposition='outside',
        textfont=dict(size=10, color='#1F2937', family="Inter"),
        hovertemplate='<b>%{x}</b><br>Costo Mensual: $%{y:,.0f}<extra></extra>'
    ))
    
    fig.update_layout(
        title="Costo Mensual Mantenimiento Inversiones No Productivas",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        xaxis=dict(
            showgrid=False, 
            tickangle=45, 
            tickfont=dict(size=10, family="Inter")
        ),
        yaxis=dict(
            showgrid=True, 
            gridcolor='#F3F4F6', 
            tickformat='$,.0f',
            tickfont=dict(size=10, family="Inter")
        ),
        font=dict(family="Inter"),
        margin=dict(l=50, r=50, t=80, b=120)
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_taxes_graphic(processed_data):
    """Create taxes graphic - Gráfica 4"""
    if 'inversiones_no_productivas' not in processed_data:
        st.warning("No non-productive investment data")
        return
    
    df_no_prod = processed_data['inversiones_no_productivas']
    
    tax_col = find_exact_column(df_no_prod, ['Impuestos'])
    name_col = find_exact_column(df_no_prod, ['Nombre del Activo'])
    
    if not tax_col or not name_col:
        st.warning(f"Tax columns not found. Available: {list(df_no_prod.columns)}")
        return
    
    df_valid = df_no_prod.copy()
    df_valid[tax_col] = pd.to_numeric(df_valid[tax_col], errors='coerce').fillna(0)
    df_valid = df_valid[df_valid[tax_col] > 0]
    
    if df_valid.empty:
        st.warning("No valid tax data")
        return
    
    nombres = df_valid[name_col].tolist()
    impuestos = df_valid[tax_col].tolist()
    # Convert annual taxes to monthly (Equation 13)
    impuestos_mensuales = [i/12 for i in impuestos]
    
    colors = ['#1E3A8A', '#3B82F6', '#60A5FA', '#93C5FD', '#DBEAFE', '#F3F4F6'][:len(nombres)]
    
    fig = go.Figure(go.Bar(
        y=nombres,
        x=impuestos_mensuales,
        orientation='h',
        marker_color=colors,
        text=[f"${i:,.0f}" if i > 0 else "" for i in impuestos_mensuales],
        textposition='auto',
        textfont=dict(color='white', size=10, family="Inter", weight="bold"),
        hovertemplate='<b>%{y}</b><br>Impuesto Mensual: $%{x:,.0f}<extra></extra>'
    ))
    
    fig.update_layout(
        title="Impuestos Mensuales Inversiones No Productivas",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        xaxis=dict(
            showgrid=True, 
            gridcolor='#F3F4F6', 
            tickformat='$,.0f',
            tickfont=dict(size=10, family="Inter")
        ),
        yaxis=dict(
            showgrid=False, 
            tickfont=dict(size=10, family="Inter")
        ),
        margin=dict(l=200, r=50, t=80, b=50),
        font=dict(family="Inter")
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_financial_investments_chart(processed_data):
    """Create financial investments chart"""
    if 'inversiones_financieras' not in processed_data:
        st.warning("No financial investment data")
        return
    
    df_fin = processed_data['inversiones_financieras']
    
    asset_class_col = find_exact_column(df_fin, ['Asset class'])
    valor_col = find_exact_column(df_fin, ['Valor (COP)'])
    
    if not asset_class_col or not valor_col:
        st.warning(f"Required columns not found. Available: {list(df_fin.columns)}")
        return
    
    df_clean = df_fin.copy()
    df_clean[asset_class_col] = df_clean[asset_class_col].astype(str).str.strip()
    df_clean[valor_col] = pd.to_numeric(df_clean[valor_col], errors='coerce').fillna(0)
    df_clean = df_clean[df_clean[valor_col] > 0]
    
    if df_clean.empty:
        st.warning("No valid financial investment data")
        return
    
    grouped = df_clean.groupby(asset_class_col)[valor_col].sum().reset_index()
    grouped = grouped[grouped[valor_col] > 0]
    
    colors_tipos = {
        'Renta fija': '#1E3A8A',
        'Renta variable': '#10B981', 
        'Alternativos': '#F59E0B'
    }
    
    colors = [colors_tipos.get(tipo.strip(), '#9CA3AF') for tipo in grouped[asset_class_col]]
    
    fig = px.pie(
        values=grouped[valor_col],
        names=grouped[asset_class_col],
        color_discrete_sequence=colors,
        hole=0.0
    )
    
    fig.update_layout(
        title="Inversiones Financieras por Tipo",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=350,
        paper_bgcolor='white',
        font=dict(family="Inter", size=12),
        margin=dict(l=10, r=10, t=60, b=10),
        showlegend=True,
        legend=dict(
            orientation="v", 
            x=1.02, 
            y=0.5,
            font=dict(size=11, family="Inter")
        )
    )
    
    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        textfont_size=11,
        textfont_family="Inter"
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_financial_sub_asset_chart(processed_data):
    """Create financial investments chart by Sub Asset class"""
    if 'inversiones_financieras' not in processed_data:
        st.warning("No financial investment data")
        return
    
    df_fin = processed_data['inversiones_financieras']
    
    sub_asset_class_col = find_exact_column(df_fin, ['Sub Asset class', 'Sub Asset Class'])
    valor_col = find_exact_column(df_fin, ['Valor (COP)'])
    
    if not sub_asset_class_col or not valor_col:
        st.warning(f"Required columns not found. Available: {list(df_fin.columns)}")
        return
    
    df_clean = df_fin.copy()
    df_clean[sub_asset_class_col] = df_clean[sub_asset_class_col].astype(str).str.strip()
    df_clean[valor_col] = pd.to_numeric(df_clean[valor_col], errors='coerce').fillna(0)
    df_clean = df_clean[df_clean[valor_col] > 0]
    
    if df_clean.empty:
        st.warning("No valid financial investment data")
        return
    
    grouped = df_clean.groupby(sub_asset_class_col)[valor_col].sum().reset_index()
    grouped = grouped[grouped[valor_col] > 0]
    
    # Extended color palette for more sub-categories
    colors_sub = ['#1E3A8A', '#3B82F6', '#60A5FA', '#10B981', '#34D399', '#F59E0B', '#FCD34D', '#8B5CF6', '#A78BFA', '#EC4899', '#F472B6', '#06B6D4']
    
    fig = px.pie(
        values=grouped[valor_col],
        names=grouped[sub_asset_class_col],
        color_discrete_sequence=colors_sub,
        hole=0.0
    )
    
    fig.update_layout(
        title="Inversiones Financieras por Sub Asset Class",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=350,
        paper_bgcolor='white',
        font=dict(family="Inter", size=12),
        margin=dict(l=10, r=10, t=60, b=10),
        showlegend=True,
        legend=dict(
            orientation="v", 
            x=1.02, 
            y=0.5,
            font=dict(size=11, family="Inter")
        )
    )
    
    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        textfont_size=11,
        textfont_family="Inter"
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_profitability_breakdown_chart(processed_data):
    """Create stacked bar chart showing Yield + Appreciation = Rentabilidad"""
    if 'inversiones_financieras' not in processed_data:
        st.warning("No financial investment data")
        return
    
    df_fin = processed_data['inversiones_financieras']
    
    nombre_col = find_exact_column(df_fin, ['Nombre del Activo'])
    yield_col = find_exact_column(df_fin, ['Yield (%)', 'Yield (%) '])
    appreciation_col = find_exact_column(df_fin, ['Apreciación Anual (%)'])
    rentabilidad_col = find_exact_column(df_fin, ['Rentabilidad (%)', 'Rentabilidad (%)'])
    
    if not all([nombre_col, yield_col, appreciation_col]):
        st.warning(f"Required columns not found. Available: {list(df_fin.columns)}")
        return
    
    df_valid = df_fin.copy()
    df_valid[yield_col] = pd.to_numeric(df_valid[yield_col], errors='coerce').fillna(0)
    df_valid[appreciation_col] = pd.to_numeric(df_valid[appreciation_col], errors='coerce').fillna(0)
    
    # Filter only rows with data
    df_valid = df_valid[(df_valid[yield_col] != 0) | (df_valid[appreciation_col] != 0)]
    
    if df_valid.empty:
        st.warning("No valid profitability data")
        return
    
    nombres = df_valid[nombre_col].tolist()
    yields = [y * 100 for y in df_valid[yield_col].tolist()]  # Convertir decimal a porcentaje
    appreciations = [a * 100 for a in df_valid[appreciation_col].tolist()]  # Convertir decimal a porcentaje
        
    fig = go.Figure()
    
    # Yield bar (bottom stack)
    fig.add_trace(go.Bar(
        name='Yield',
        x=nombres,
        y=yields,
        marker_color='#3B82F6',
        text=[f"{y:.1f}%" if y > 0 else "" for y in yields],
        textposition='inside',
        textfont=dict(size=10, color='white', family="Inter"),
        hovertemplate='<b>%{x}</b><br>Yield: %{y:.2f}%<extra></extra>'
    ))
    
    # Appreciation bar (top stack)
    fig.add_trace(go.Bar(
        name='Apreciación',
        x=nombres,
        y=appreciations,
        marker_color='#10B981',
        text=[f"{a:.1f}%" if a > 0 else "" for a in appreciations],
        textposition='inside',
        textfont=dict(size=10, color='white', family="Inter"),
        hovertemplate='<b>%{x}</b><br>Apreciación: %{y:.2f}%<extra></extra>'
    ))
    
    fig.update_layout(
        title="Composición de Rentabilidad por Activo Financiero",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=450,
        paper_bgcolor='white',
        plot_bgcolor='white',
        barmode='stack',
        xaxis=dict(
            title="Instrumento Financiero",
            showgrid=False,
            tickangle=45,
            tickfont=dict(size=10, family="Inter"),
            title_font=dict(size=12, family="Inter")
        ),
        yaxis=dict(
            title="Rentabilidad (%)",
            showgrid=True,
            gridcolor='#F3F4F6',
            ticksuffix='%',
            tickfont=dict(size=12, family="Inter"),
            title_font=dict(size=14, family="Inter")
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5,
            font=dict(size=12, family="Inter")
        ),
        font=dict(family="Inter"),
        margin=dict(l=80, r=50, t=80, b=150)
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_currency_chart(processed_data):
    """Create currency distribution chart"""
    if 'inversiones_financieras' not in processed_data:
        st.warning("No financial investment data")
        return
    
    df_fin = processed_data['inversiones_financieras']
    
    asset_class_col = find_exact_column(df_fin, ['Asset class'])
    moneda_col = find_exact_column(df_fin, ['Moneda (Lista)'])
    valor_col = find_exact_column(df_fin, ['Valor (COP)'])
    
    if not all([asset_class_col, moneda_col, valor_col]):
        missing_cols = []
        if not asset_class_col: missing_cols.append('Asset class')
        if not moneda_col: missing_cols.append('Moneda (Lista)')
        if not valor_col: missing_cols.append('Valor (COP)')
        st.warning(f"Missing columns: {missing_cols}. Available: {list(df_fin.columns)}")
        return
    
    df_clean = df_fin.copy()
    df_clean[asset_class_col] = df_clean[asset_class_col].astype(str).str.strip()
    df_clean[moneda_col] = df_clean[moneda_col].astype(str).str.strip()
    df_clean[valor_col] = pd.to_numeric(df_clean[valor_col], errors='coerce').fillna(0)
    df_clean = df_clean[df_clean[valor_col] > 0]
    
    if df_clean.empty:
        st.warning("No valid data for currency chart")
        return
    
    grouped = df_clean.groupby([asset_class_col, moneda_col])[valor_col].sum().reset_index()
    
    fig = go.Figure()
    
    tipos_inversion = grouped[asset_class_col].unique()
    monedas = grouped[moneda_col].unique()
    
    colors_monedas = {
        'COP': '#1E3A8A',
        'USD': '#10B981', 
        'EUR': '#F59E0B',
        'GBP': '#8B5CF6',
        'JPY': '#EF4444',
        'CAD': '#06B6D4'
    }
    
    for moneda in monedas:
        moneda_data = grouped[grouped[moneda_col] == moneda]
        
        x_values = []
        y_values = []
        
        for tipo in tipos_inversion:
            tipo_data = moneda_data[moneda_data[asset_class_col] == tipo]
            x_values.append(tipo)
            if len(tipo_data) > 0:
                y_values.append(tipo_data[valor_col].iloc[0])
            else:
                y_values.append(0)
        
        fig.add_trace(go.Bar(
            name=moneda,
            x=x_values,
            y=y_values,
            marker_color=colors_monedas.get(moneda, '#9CA3AF'),
            text=[f"${v:,.0f}" if v > 0 else "" for v in y_values],
            textposition='inside',
            textfont=dict(size=10, color='white', family="Inter"),
            hovertemplate=f'<b>{moneda}</b><br>%{{x}}<br>Valor: $%{{y:,.0f}}<extra></extra>'
        ))
    
    fig.update_layout(
        title="Distribución de Inversiones por Tipo y Moneda",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=450,
        paper_bgcolor='white',
        plot_bgcolor='white',
        barmode='stack',
        xaxis=dict(
            title="Tipo de Inversión",
            tickfont=dict(size=12, family="Inter"),
            title_font=dict(size=14, family="Inter")
        ),
        yaxis=dict(
            title="Valor (COP)",
            tickformat='$,.0f',
            tickfont=dict(size=12, family="Inter"),
            title_font=dict(size=14, family="Inter")
        ),
        legend=dict(
            title="Moneda",
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02,
            font=dict(size=12, family="Inter")
        ),
        font=dict(family="Inter"),
        margin=dict(l=80, r=150, t=80, b=80)
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_currency_pie_chart(processed_data):
    """Create pie chart showing total value by currency only (no asset type breakdown)"""
    if 'inversiones_financieras' not in processed_data:
        st.warning("No financial investment data")
        return
    
    df_fin = processed_data['inversiones_financieras']
    
    moneda_col = find_exact_column(df_fin, ['Moneda (Lista)'])
    valor_col = find_exact_column(df_fin, ['Valor (COP)'])
    
    if not all([moneda_col, valor_col]):
        missing_cols = []
        if not moneda_col: missing_cols.append('Moneda (Lista)')
        if not valor_col: missing_cols.append('Valor (COP)')
        st.warning(f"Missing columns: {missing_cols}. Available: {list(df_fin.columns)}")
        return
    
    df_clean = df_fin.copy()
    df_clean[moneda_col] = df_clean[moneda_col].astype(str).str.strip()
    df_clean[valor_col] = pd.to_numeric(df_clean[valor_col], errors='coerce').fillna(0)
    df_clean = df_clean[df_clean[valor_col] > 0]
    
    if df_clean.empty:
        st.warning("No valid data for currency pie chart")
        return
    
    # Group by currency only - sum all asset types
    grouped = df_clean.groupby(moneda_col)[valor_col].sum().reset_index()
    grouped = grouped[grouped[valor_col] > 0].sort_values(valor_col, ascending=False)
    
    colors_monedas = {
        'COP': '#1E3A8A',
        'USD': '#10B981', 
        'EUR': '#F59E0B',
        'GBP': '#8B5CF6',
        'JPY': '#EF4444',
        'CAD': '#06B6D4'
    }
    
    colors = [colors_monedas.get(moneda, '#9CA3AF') for moneda in grouped[moneda_col]]
    
    fig = px.pie(
        values=grouped[valor_col],
        names=grouped[moneda_col],
        color_discrete_sequence=colors,
        hole=0.0
    )
    
    fig.update_layout(
        title="Distribución Total por Moneda",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=400,
        paper_bgcolor='white',
        font=dict(family="Inter", size=12),
        margin=dict(l=10, r=10, t=60, b=10),
        showlegend=True,
        legend=dict(
            orientation="v", 
            x=1.02, 
            y=0.5,
            font=dict(size=12, family="Inter")
        )
    )
    
    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        textfont_size=12,
        textfont_family="Inter",
        hovertemplate='<b>%{label}</b><br>Valor: $%{value:,.0f}<br>%{percent}<extra></extra>'
    )
    
    st.plotly_chart(fig, use_container_width=True)

def display_cash_flow_table(flow_analysis):
    """Display comprehensive cash flow table according to manual methodology"""
    if not flow_analysis:
        st.error("No cash flow data to display")
        return
    
    try:
        ingresos = flow_analysis.get('ingresos', {})
        gastos_p1 = flow_analysis.get('gastos_p1', {})
        gastos_p2 = flow_analysis.get('gastos_p2', {})
        inversiones = flow_analysis.get('inversiones', {})
        impuestos = flow_analysis.get('impuestos', {})
        resumen = flow_analysis.get('resumen', {})
        porcentajes = resumen.get('porcentajes', {})
        
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #1E3A8A; font-weight: 700; font-size: 1.5rem; margin-bottom: 0.5rem;">
                ANÁLISIS DE FLUJO DE EFECTIVO REQUERIDO
            </h2>
            <p style="color: #6B7280; font-size: 0.875rem;">
                Metodología Proaltus de Priorización de Gastos (Mensual)
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        data = []
        row_styles = []
        
        # Income section
        data.append(['Ingreso', f"${safe_float(ingresos.get('total', 0)):,.0f}", '100%'])
        row_styles.append('highlight')
        
        data.append(['  Ingreso Salarial', f"${safe_float(ingresos.get('ingreso_salarial', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Ingresos Pasivos', f"${safe_float(ingresos.get('ingresos_pasivos', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        # Priority 1 expenses (GP1) - Equation 9
        data.append(['Gastos Prioridad 1 (GP1)', f"${safe_float(gastos_p1.get('total', 0)):,.0f}", f"{safe_float(porcentajes.get('gastos_p1', 0)):.0f}%"])
        row_styles.append('highlight')
        
        data.append(['  Gastos Esenciales (servicios, colegio, mercado, salud)', f"${safe_float(gastos_p1.get('gastos_esenciales', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Gastos Operativos (empleada, gasolina)', f"${safe_float(gastos_p1.get('gastos_operativos', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Mantenimiento Inversiones (anual/12)', f"${safe_float(gastos_p1.get('relacionado_inversiones', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        # Priority 2 expenses (GP2) - Equation 10
        data.append(['Gastos Prioridad 2 (GP2)', f"${safe_float(gastos_p2.get('total', 0)):,.0f}", f"{safe_float(porcentajes.get('gastos_p2', 0)):.0f}%"])
        row_styles.append('highlight')
        
        data.append(['  Gastos Varios (comidas, deportes, ropa, clubes)', f"${safe_float(gastos_p2.get('gastos_varios', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Viajes', f"${safe_float(gastos_p2.get('viajes', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Lujo (joyas, arte)', f"${safe_float(gastos_p2.get('lujo', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        # Investments (INV)
        data.append(['Inversiones (INV)', f"${safe_float(inversiones.get('total', 0)):,.0f}", f"{safe_float(porcentajes.get('inversiones', 0)):.0f}%"])
        row_styles.append('normal')
        
        data.append(['  Aporte a Pensión Voluntaria', f"${safe_float(inversiones.get('pension_voluntaria', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Compromiso Proyecto Inmobiliarios', f"${safe_float(inversiones.get('proyecto_inmobiliarios', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        # Taxes (IMP) - Equation 13
        data.append(['Impuestos (IMP)', f"${safe_float(impuestos.get('total', 0)):,.0f}", f"{safe_float(porcentajes.get('impuestos', 0)):.0f}%"])
        row_styles.append('highlight')
        
        data.append(['  Impuestos Inversiones (anual/12)', f"${safe_float(impuestos.get('impuestos_inversiones', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        data.append(['  Provisión Tributaria (renta, patrimonio)', f"${safe_float(impuestos.get('provision_impuestos', 0)):,.0f}", ''])
        row_styles.append('normal')
        
        # Totals - Equations 2 and 4
        data.append(['TOTAL EGRESOS (GP1+GP2+INV+IMP)', f"${safe_float(resumen.get('total_egresos', 0)):,.0f}", ''])
        row_styles.append('highlight')
        
        data.append(['Flujo de Efectivo Neto (FCN)', f"${safe_float(resumen.get('resultado_neto', 0)):,.0f}", f"{safe_float(porcentajes.get('resultado_neto', 0)):.0f}%"])
        row_styles.append('highlight')
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=['FLUJO REQUERIDO (Mensual)', 'VALOR $', '%'])
        
        # Apply styling
        def highlight_rows(row):
            idx = row.name
            if idx < len(row_styles):
                if row_styles[idx] == 'highlight':
                    return ['background-color: #DBEAFE; font-weight: bold; border-left: 4px solid #1E3A8A;'] * len(row)
                else:
                    return [''] * len(row)
            return [''] * len(row)
        
        styled_df = df.style.apply(highlight_rows, axis=1)
        
        st.dataframe(
            styled_df,
            use_container_width=True,
            hide_index=True
        )
            
    except Exception as e:
        st.error(f"Error displaying cash flow table: {str(e)}")

def create_return_graphic():
    """Create expected returns chart - Gráfica 10"""
    categories = [
        'Empresas',
        'Inversiones No productivas\n(Terrenos, Aviones, bienes inmuebles...)',
        'Inversiones Productivas', 
        'Inversiones Financieras'
    ]
    
    # Reference values according to manual
    returns = [3, 2, 8, 7]
    colors = ['#1E3A8A', '#1E3A8A', '#9CA3AF', '#9CA3AF']
    
    fig = go.Figure(go.Bar(
        x=categories,
        y=returns,
        marker_color=colors,
        text=[f"{r}%" for r in returns],
        textposition='outside',
        textfont=dict(size=16, color='#1F2937', family="Inter", weight="bold"),
        hovertemplate='<b>%{x}</b><br>Rendimiento: %{y}%<extra></extra>'
    ))
    
    fig.update_layout(
        title="Rendimiento Esperado por Categoría de Activos",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=450,
        paper_bgcolor='white',
        plot_bgcolor='white',
        xaxis=dict(
            showgrid=False, 
            tickangle=45, 
            tickfont=dict(size=11, color='#1F2937', family="Inter")
        ),
        yaxis=dict(
            showgrid=True, 
            gridcolor='#F3F4F6',
            ticksuffix='%',
            range=[0, 10],
            tickfont=dict(size=12, family="Inter")
        ),
        font=dict(family="Inter"),
        margin=dict(l=50, r=50, t=80, b=150)
    )
    
    st.plotly_chart(fig, use_container_width=True)

def create_geographic_distribution_map(processed_data):
    """Create interactive map showing asset distribution by geography"""
    
    geographic_data = {}
    
    # Collect data from all investment types
    sheets_config = {
        'inversiones_productivas': {'valor': 'Valor (COP)', 'geo': 'Geografia', 'nombre': 'Nombre del Proyecto'},
        'inversiones_no_productivas': {'valor': 'Valor (COP)', 'geo': 'Geografia ', 'nombre': 'Nombre del Activo'},
        'inversiones_financieras': {'valor': 'Valor (COP)', 'geo': 'Geografia', 'nombre': 'Nombre del Activo'}
    }
    
    for sheet_key, cols in sheets_config.items():
        if sheet_key not in processed_data:
            continue
            
        df = processed_data[sheet_key]
        
        valor_col = find_exact_column(df, [cols['valor']])
        geo_col = find_exact_column(df, [cols['geo'], 'Geografia', 'Geografia ', 'Geography'])
        
        if not valor_col or not geo_col:
            continue
        
        # Mapping de códigos a nombres completos
        country_mapping = {
            'COL': 'Colombia',
            'USA': 'United States',
            'ESP': 'Spain',
            'MEX': 'Mexico',
            'BRA': 'Brazil',
            'ARG': 'Argentina',
            'CHI': 'Chile',
            'PER': 'Peru',
            'GBR': 'United Kingdom',
            'UK': 'United Kingdom',
            'CAN': 'Canada',
            'JPN': 'Japan',
            'CHN': 'China'
        }
        
        for _, row in df.iterrows():
            geografia_raw = str(row[geo_col]).strip().upper() if pd.notna(row[geo_col]) else 'No especificado'
            # Convertir código a nombre completo
            geografia = country_mapping.get(geografia_raw, geografia_raw)
            valor = safe_float(row[valor_col])
            
            if valor > 0 and geografia and geografia.lower() not in ['nan', '', 'none', 'no especificado']:
                if geografia in geographic_data:
                    geographic_data[geografia]['valor'] += valor
                    geographic_data[geografia]['cantidad'] += 1
                else:
                    geographic_data[geografia] = {'valor': valor, 'cantidad': 1}
    
    if not geographic_data:
        st.warning("No se encontraron datos geográficos para mostrar en el mapa")
        return
    
    # Create DataFrame for map
    df_map = pd.DataFrame([
        {
            'Ubicación': loc,
            'Valor (M COP)': data['valor'] / 1_000_000,
            'Cantidad de Activos': data['cantidad']
        }
        for loc, data in geographic_data.items()
    ]).sort_values('Valor (M COP)', ascending=False)
    
    # Corporate color scale matching dashboard
    custom_colors = ["#DBEAFE", "#93C5FD", "#60A5FA", "#3B82F6", "#1E3A8A"]
    
    fig = px.choropleth(
        df_map,
        locations="Ubicación",
        locationmode="country names",
        color="Valor (M COP)",
        hover_name="Ubicación",
        hover_data={
            'Valor (M COP)': ':,.0f',
            'Cantidad de Activos': ':,',
            'Ubicación': False
        },
        color_continuous_scale=custom_colors,
        projection="natural earth"
    )
    
    fig.update_layout(
        title=dict(
            text="Distribución Geográfica de Activos",
            font=dict(size=18, color='#1F2937', family="Inter"),
            x=0.5,
            xanchor='center'
        ),
        geo=dict(
            showframe=False,
            showcoastlines=True,
            coastlinecolor="#E5E7EB",
            showland=True,
            landcolor="#F3F4F6",
            bgcolor="white",
            projection_type="natural earth",
            showcountries=True,
            countrycolor="#E5E7EB"
        ),
        coloraxis_colorbar=dict(
            title=dict(text="Valor (M COP)", font=dict(family="Inter", size=12)),
            ticks="outside",
            showticklabels=True,
            tickfont=dict(family="Inter", size=10)
        ),
        height=500,
        paper_bgcolor='white',
        font=dict(family="Inter"),
        margin={"r":20,"t":60,"l":20,"b":20}
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Summary table
    with st.expander("Ver detalle por ubicación"):
        df_display = df_map.copy()
        df_display['Valor (M COP)'] = df_display['Valor (M COP)'].apply(lambda x: f"${x:,.0f}M")
        st.dataframe(df_display, use_container_width=True, hide_index=True)
def create_cost_comparison_chart(processed_data):
    """Create cost comparison chart: Current Management vs Proaltus"""
    
    if 'inversiones_financieras' not in processed_data:
        st.warning("No financial investment data available")
        return
    
    df_fin = processed_data['inversiones_financieras']
    
    # Find columns - NOMBRES CORRECTOS DEL EXCEL
    nombre_col = find_exact_column(df_fin, ['Nombre del Activo'])
    valor_col = find_exact_column(df_fin, ['Valor (COP)'])
    fee_actual_pct_col = find_exact_column(df_fin, [
        'Management Fee Actual (%)',
        'Management Fee Actual(%)',
        'Management Fee Actual (%) '
    ])
    fee_proaltus_pct_col = find_exact_column(df_fin, [
        'Costo Proaltus (%)',
        'Costo Proaltus(%)',
        'Costo Proaltus (%) '
    ])
    
    if not all([nombre_col, valor_col, fee_actual_pct_col, fee_proaltus_pct_col]):
        missing = []
        if not nombre_col: missing.append('Nombre del Activo')
        if not valor_col: missing.append('Valor (COP)')
        if not fee_actual_pct_col: missing.append('Management Fee Actual (%)')
        if not fee_proaltus_pct_col: missing.append('Costo Proaltus (%)')
        st.warning(f"Required columns not found: {missing}. Available: {list(df_fin.columns)}")
        return
    
    # Calculate totals
    df_valid = df_fin.copy()
    df_valid[valor_col] = pd.to_numeric(df_valid[valor_col], errors='coerce').fillna(0)
    df_valid[fee_actual_pct_col] = pd.to_numeric(df_valid[fee_actual_pct_col], errors='coerce').fillna(0)
    df_valid[fee_proaltus_pct_col] = pd.to_numeric(df_valid[fee_proaltus_pct_col], errors='coerce').fillna(0)
    
    # Calcular costos totales anuales
    total_actual = 0
    total_proaltus = 0
    
    for _, row in df_valid.iterrows():
        valor = safe_float(row[valor_col])
        fee_actual_pct = safe_float(row[fee_actual_pct_col])
        fee_proaltus_pct = safe_float(row[fee_proaltus_pct_col])
        
        # Costo anual = Valor × Fee% / 100
        costo_actual_anual = valor * (fee_actual_pct / 100)
        costo_proaltus_anual = valor * (fee_proaltus_pct / 100)
        
        total_actual += costo_actual_anual
        total_proaltus += costo_proaltus_anual
    
    ahorro_anual = total_actual - total_proaltus
    ahorro_mensual = ahorro_anual / 12
    ahorro_porcentaje = (ahorro_anual / total_actual * 100) if total_actual > 0 else 0
    
    # Create comparison chart
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        name='Gestor Actual',
        x=['Costo Anual Management'],
        y=[total_actual],
        marker_color='#DC2626',
        text=[f"${total_actual:,.0f}"],
        textposition='auto',
        textfont=dict(color='white', size=14, family="Inter", weight="bold"),
        hovertemplate='<b>Gestor Actual</b><br>Costo: $%{y:,.0f}<extra></extra>'
    ))
    
    fig.add_trace(go.Bar(
        name='Proaltus',
        x=['Costo Anual Management'],
        y=[total_proaltus],
        marker_color='#059669',
        text=[f"${total_proaltus:,.0f}"],
        textposition='auto',
        textfont=dict(color='white', size=14, family="Inter", weight="bold"),
        hovertemplate='<b>Proaltus</b><br>Costo: $%{y:,.0f}<extra></extra>'
    ))
    
    fig.update_layout(
        title="Comparación de Costos de Gestión - Actual vs Proaltus",
        title_font_size=16,
        title_font_color='#1F2937',
        title_font_family="Inter",
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        barmode='group',
        xaxis=dict(
            showgrid=False,
            tickfont=dict(size=12, family="Inter")
        ),
        yaxis=dict(
            title="Costo Anual (COP)",
            showgrid=True,
            gridcolor='#F3F4F6',
            tickformat='$,.0f',
            tickfont=dict(size=12, family="Inter"),
            title_font=dict(size=14, family="Inter")
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5,
            font=dict(size=12, family="Inter")
        ),
        font=dict(family="Inter"),
        margin=dict(l=80, r=50, t=80, b=50)
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Savings summary
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Ahorro Anual</div>
            <div class="kpi-value" style="color: #059669;">${ahorro_anual:,.0f}</div>
            <div class="kpi-meta">{ahorro_porcentaje:.1f}% de reducción</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Ahorro Mensual</div>
            <div class="kpi-value" style="color: #059669;">${ahorro_mensual:,.0f}</div>
            <div class="kpi-meta">Promedio por mes</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Valor Agregado</div>
            <div class="kpi-value" style="color: #1E3A8A;">${ahorro_anual * 5:,.0f}</div>
            <div class="kpi-meta">Ahorro proyectado a 5 años</div>
        </div>
        """, unsafe_allow_html=True)

# PDF GENERATION
def save_chart_as_image(fig, filename, width=800, height=500):
    """Save Plotly figure as image file"""
    try:
        import tempfile
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        fig.write_image(filepath, width=width, height=height, format='png')
        return filepath
    except Exception as e:
        st.warning(f"Could not save chart {filename}: {str(e)}")
        return None
    
def generate_pdf_report(flow_analysis, kpis, processed_data):
    """Generate comprehensive PDF report with ALL charts and data"""
    if not PDF_AVAILABLE:
        st.error("PDF libraries not available.")
        return None
    
    try:
        import tempfile
        import os
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        chart_images = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            textColor=colors.HexColor('#1E3A8A'),
            alignment=1,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            textColor=colors.HexColor('#1E3A8A'),
            fontName='Helvetica-Bold'
        )
        
        # ===== TITLE PAGE =====
        story.append(Paragraph("PROALTUS - RADIOGRAFÍA FINANCIERA", title_style))
        story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # ===== RESUMEN EJECUTIVO =====
        story.append(Paragraph("RESUMEN EJECUTIVO", subtitle_style))
        
        kpi_data = [
            ['Métrica', 'Valor'],
            ['Patrimonio Total', f"${safe_float(kpis.get('total_patrimony', 0)):,.0f}"],
            ['Ingresos Mensuales', f"${safe_float(kpis.get('total_income', 0)):,.0f}"],
            ['Flujo Efectivo Neto (FCN)', f"${safe_float(kpis.get('net_flow', 0)):,.0f}"],
            ['Tasa de Ahorro (TA)', f"{safe_float(kpis.get('savings_rate', 0)):.1f}%"],
            ['Número de Activos', str(int(kpis.get('asset_count', 0)))]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 30))
        
        # ===== ANÁLISIS DE FLUJO DE EFECTIVO REQUERIDO (RIGHT AFTER SUMMARY) =====
        if flow_analysis:
            story.append(Paragraph("ANÁLISIS DE FLUJO DE EFECTIVO REQUERIDO", subtitle_style))
            story.append(Paragraph("Metodología Proaltus de Priorización de Gastos (Mensual)", styles['Normal']))
            story.append(Spacer(1, 10))
            
            ingresos = flow_analysis['ingresos']
            gastos_p1 = flow_analysis['gastos_p1']
            gastos_p2 = flow_analysis['gastos_p2']
            inversiones = flow_analysis['inversiones']
            impuestos = flow_analysis['impuestos']
            resumen = flow_analysis['resumen']
            porcentajes = resumen['porcentajes']
            
            flow_data = [
                ['FLUJO REQUERIDO (Mensual)', 'VALOR $', '%'],
                ['Ingreso', f"${ingresos['total']:,.0f}", '100%'],
                ['  Ingreso Salarial', f"${ingresos['ingreso_salarial']:,.0f}", ''],
                ['  Ingresos Pasivos', f"${ingresos['ingresos_pasivos']:,.0f}", ''],
                ['', '', ''],
                ['Gastos Prioridad 1 (GP1)', f"${gastos_p1['total']:,.0f}", f"{porcentajes['gastos_p1']:.0f}%"],
                ['  Gastos Esenciales', f"${gastos_p1['gastos_esenciales']:,.0f}", ''],
                ['  Gastos Operativos', f"${gastos_p1['gastos_operativos']:,.0f}", ''],
                ['  Mantenimiento Inversiones', f"${gastos_p1['relacionado_inversiones']:,.0f}", ''],
                ['', '', ''],
                ['Gastos Prioridad 2 (GP2)', f"${gastos_p2['total']:,.0f}", f"{porcentajes['gastos_p2']:.0f}%"],
                ['  Gastos Varios', f"${gastos_p2['gastos_varios']:,.0f}", ''],
                ['  Viajes', f"${gastos_p2['viajes']:,.0f}", ''],
                ['  Lujo', f"${gastos_p2['lujo']:,.0f}", ''],
                ['', '', ''],
                ['Inversiones (INV)', f"${inversiones['total']:,.0f}", f"{porcentajes['inversiones']:.0f}%"],
                ['  Aporte Pensión Voluntaria', f"${inversiones['pension_voluntaria']:,.0f}", ''],
                ['  Proyecto Inmobiliarios', f"${inversiones['proyecto_inmobiliarios']:,.0f}", ''],
                ['', '', ''],
                ['Impuestos (IMP)', f"${impuestos['total']:,.0f}", f"{porcentajes['impuestos']:.0f}%"],
                ['  Impuestos Inversiones', f"${impuestos['impuestos_inversiones']:,.0f}", ''],
                ['  Provisión Tributaria', f"${impuestos['provision_impuestos']:,.0f}", ''],
                ['', '', ''],
                ['TOTAL EGRESOS', f"${resumen['total_egresos']:,.0f}", ''],
                ['Flujo Efectivo Neto (FCN)', f"${resumen['resultado_neto']:,.0f}", f"{porcentajes['resultado_neto']:.0f}%"]
            ]
            
            flow_table = Table(flow_data, colWidths=[3*inch, 1.5*inch, 0.75*inch])
            flow_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#DBEAFE')),
                ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#DBEAFE')),
                ('BACKGROUND', (0, 10), (-1, 10), colors.HexColor('#DBEAFE')),
                ('BACKGROUND', (0, 15), (-1, 15), colors.HexColor('#DBEAFE')),
                ('BACKGROUND', (0, 19), (-1, 19), colors.HexColor('#DBEAFE')),
                ('BACKGROUND', (0, 23), (-1, -1), colors.HexColor('#DBEAFE')),
                ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
                ('FONTNAME', (0, 5), (0, 5), 'Helvetica-Bold'),
                ('FONTNAME', (0, 10), (0, 10), 'Helvetica-Bold'),
                ('FONTNAME', (0, 15), (0, 15), 'Helvetica-Bold'),
                ('FONTNAME', (0, 19), (0, 19), 'Helvetica-Bold'),
                ('FONTNAME', (0, 23), (0, -1), 'Helvetica-Bold'),
            ]))
            
            story.append(flow_table)
            story.append(Spacer(1, 30))
        
        # ===== NOW ADD ALL CHARTS =====
        story.append(Paragraph("ANÁLISIS GRÁFICO", subtitle_style))
        story.append(Spacer(1, 10))
        
        try:
            # 1. CASH FLOW CHART
            if flow_analysis:
                story.append(Paragraph("Gráfica 1: Estructura de Ingresos", subtitle_style))
                
                ingresos = flow_analysis['ingresos']
                fig = go.Figure()
                
                fig.add_trace(go.Bar(
                    y=['Ingreso Total'],
                    x=[ingresos['total']],
                    orientation='h',
                    marker_color='#1E3A8A',
                    text=[f"${ingresos['total']:,.0f}"],
                    textposition='inside',
                    textfont=dict(color='white', size=12)
                ))
                
                fig.add_trace(go.Bar(
                    y=['Ingreso Salarial'],
                    x=[ingresos['ingreso_salarial']],
                    orientation='h',
                    marker_color='#F3F4F6',
                    text=[f"${ingresos['ingreso_salarial']:,.0f}"],
                    textposition='inside',
                    textfont=dict(color='#1F2937', size=12)
                ))
                
                fig.add_trace(go.Bar(
                    y=['Ingresos Pasivos'],
                    x=[ingresos['ingresos_pasivos']],
                    orientation='h',
                    marker_color='#60A5FA',
                    text=[f"${ingresos['ingresos_pasivos']:,.0f}"],
                    textposition='inside',
                    textfont=dict(color='white', size=12)
                ))
                
                fig.update_layout(
                    height=300,
                    showlegend=False,
                    paper_bgcolor='white',
                    plot_bgcolor='white',
                    xaxis=dict(showgrid=False, showticklabels=False),
                    yaxis=dict(categoryorder='array', categoryarray=['Ingresos Pasivos', 'Ingreso Salarial', 'Ingreso Total']),
                    margin=dict(l=100, r=50, t=20, b=50)
                )
                
                img_path = save_chart_as_image(fig, 'cash_flow.png', width=700, height=300)
                if img_path:
                    chart_images.append(img_path)
                    story.append(Image(img_path, width=5.5*inch, height=2.4*inch))
                    story.append(Spacer(1, 20))
                
                # 2. EXPENSE STRUCTURE CHART
            if flow_analysis:
                story.append(Paragraph("Gráfica 2: Estructura de Gastos", subtitle_style))
                
                gastos_p1 = flow_analysis['gastos_p1']
                gastos_p2 = flow_analysis['gastos_p2']
                inversiones = flow_analysis['inversiones']
                impuestos = flow_analysis['impuestos']
                resumen = flow_analysis['resumen']
                
                categories_gastos = ['GP1', 'GP2', 'INV', 'IMP']
                values_gastos = [
                    gastos_p1['total'],
                    gastos_p2['total'],
                    inversiones['total'],
                    impuestos['total']
                ]
                colors_gastos = ['#1E3A8A', '#3B82F6', '#60A5FA', '#93C5FD']
                
                fig_gastos = go.Figure(go.Bar(
                    x=categories_gastos,
                    y=values_gastos,
                    marker_color=colors_gastos,
                    text=[f"${v:,.0f}" for v in values_gastos],
                    textposition='outside',
                    textfont=dict(size=11, color='#1F2937')
                ))
                
                fig_gastos.update_layout(
                    height=350,
                    paper_bgcolor='white',
                    plot_bgcolor='white',
                    xaxis=dict(
                        showgrid=False,
                        title="Categoría de Gasto",
                        tickfont=dict(size=12)
                    ),
                    yaxis=dict(
                        showgrid=True, 
                        gridcolor='#F3F4F6',
                        title="Monto (COP)",
                        tickformat='$,.0f'
                    ),
                    margin=dict(l=80, r=50, t=20, b=80)
                )
                
                img_path = save_chart_as_image(fig_gastos, 'expense_structure.png', width=700, height=350)
                if img_path:
                    chart_images.append(img_path)
                    story.append(Image(img_path, width=5.5*inch, height=2.8*inch))
                    story.append(Spacer(1, 20))
            
            # 3. PATRIMONY DISTRIBUTION PIE
            story.append(Paragraph("Gráfica: Distribución del Patrimonio", subtitle_style))
            
            categories = ['Empresas', 'Inv. Productivas', 'Inv. No Productivas', 'Inv. Financieras']
            values = [
                safe_float(kpis.get('total_companies', 0)),
                safe_float(kpis.get('total_productive', 0)),
                safe_float(kpis.get('total_non_productive', 0)),
                safe_float(kpis.get('total_financial', 0))
            ]
            
            fig_patrimony = px.pie(
                values=values,
                names=categories,
                color_discrete_sequence=['#1E3A8A', '#10B981', '#F59E0B', '#8B5CF6'],
                hole=0.0
            )
            
            fig_patrimony.update_layout(
                height=350,
                paper_bgcolor='white',
                margin=dict(l=20, r=20, t=20, b=20),
                showlegend=True,
                legend=dict(x=0.7, y=0.5)
            )
            
            fig_patrimony.update_traces(textposition='inside', textinfo='percent+label', textfont_size=10)
            
            img_path = save_chart_as_image(fig_patrimony, 'patrimony.png', width=700, height=350)
            if img_path:
                chart_images.append(img_path)
                story.append(Image(img_path, width=5.5*inch, height=2.8*inch))
                story.append(Spacer(1, 20))
            
            # 3. MAINTENANCE COSTS
            if 'inversiones_no_productivas' in processed_data:
                story.append(Paragraph("Gráfica 3: Costos de Mantenimiento Mensual", subtitle_style))
                
                df_no_prod = processed_data['inversiones_no_productivas']
                name_col = find_exact_column(df_no_prod, ['Nombre del Activo'])
                costo_mant_col = find_exact_column(df_no_prod, ['Costo mantenimiento'])
                if name_col and costo_mant_col:
                    df_valid = df_no_prod.copy()
                    df_valid[costo_mant_col] = pd.to_numeric(df_valid[costo_mant_col], errors='coerce').fillna(0)
                    
                    nombres = []
                    costos = []
                    
                    for _, row in df_valid.iterrows():
                        costo_anual = safe_float(row[costo_mant_col])
                        costo_mensual = costo_anual / 12
                        
                        if costo_mensual > 0:
                            nombres.append(str(row[name_col]))
                            costos.append(costo_mensual)

                                    
                    if costos:
                        fig_maint = go.Figure(go.Bar(
                            x=nombres,
                            y=costos,
                            marker_color='#3B82F6',
                            text=[f"${c:,.0f}" for c in costos],
                            textposition='outside'
                        ))
                        
                        fig_maint.update_layout(
                            height=350,
                            paper_bgcolor='white',
                            plot_bgcolor='white',
                            xaxis=dict(tickangle=45),
                            yaxis=dict(showgrid=True, gridcolor='#F3F4F6'),
                            margin=dict(l=50, r=50, t=20, b=120)
                        )
                        
                        img_path = save_chart_as_image(fig_maint, 'maintenance.png', width=700, height=350)
                        if img_path:
                            chart_images.append(img_path)
                            story.append(Image(img_path, width=5.5*inch, height=2.8*inch))
                            story.append(Spacer(1, 20))
            
            # 4. FINANCIAL INVESTMENTS BY ASSET CLASS
            if 'inversiones_financieras' in processed_data:
                story.append(Paragraph("Inversiones Financieras por Asset Class", subtitle_style))
                
                df_fin = processed_data['inversiones_financieras']
                asset_class_col = find_exact_column(df_fin, ['Asset class'])
                valor_col = find_exact_column(df_fin, ['Valor (COP)'])
                
                if asset_class_col and valor_col:
                    df_clean = df_fin.copy()
                    df_clean[valor_col] = pd.to_numeric(df_clean[valor_col], errors='coerce').fillna(0)
                    grouped = df_clean.groupby(asset_class_col)[valor_col].sum().reset_index()
                    
                    fig_fin = px.pie(
                        values=grouped[valor_col],
                        names=grouped[asset_class_col],
                        color_discrete_sequence=['#1E3A8A', '#10B981', '#F59E0B'],
                        hole=0.0
                    )
                    
                    fig_fin.update_layout(
                        height=300,
                        paper_bgcolor='white',
                        margin=dict(l=20, r=20, t=20, b=20),
                        showlegend=True
                    )
                    
                    fig_fin.update_traces(textposition='inside', textinfo='percent+label')
                    
                    img_path = save_chart_as_image(fig_fin, 'financial_assets.png', width=700, height=300)
                    if img_path:
                        chart_images.append(img_path)
                        story.append(Image(img_path, width=5.5*inch, height=2.4*inch))
                        story.append(Spacer(1, 20))
            
            # 5. EXPECTED RETURNS
            story.append(Paragraph("Gráfica 10: Rendimiento Esperado por Tipo de Activo", subtitle_style))
            
            categories_ret = [
                'Empresas',
                'Inv. No Productivas',
                'Inv. Productivas',
                'Inv. Financieras'
            ]
            returns = [3, 2, 8, 7]
            
            fig_returns = go.Figure(go.Bar(
                x=categories_ret,
                y=returns,
                marker_color=['#1E3A8A', '#1E3A8A', '#9CA3AF', '#9CA3AF'],
                text=[f"{r}%" for r in returns],
                textposition='outside'
            ))
            
            fig_returns.update_layout(
                height=350,
                paper_bgcolor='white',
                plot_bgcolor='white',
                xaxis=dict(tickangle=45),
                yaxis=dict(ticksuffix='%', range=[0, 10]),
                margin=dict(l=50, r=50, t=20, b=100)
            )
            
            img_path = save_chart_as_image(fig_returns, 'returns.png', width=700, height=350)
            if img_path:
                chart_images.append(img_path)
                story.append(Image(img_path, width=5.5*inch, height=2.8*inch))
                story.append(Spacer(1, 20))
                
        except Exception as e:
            st.warning(f"Could not generate some chart images: {str(e)}")
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            f"Generado por Proaltus Dashboard v4.0 - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)
        ))
        
        doc.build(story)
        
        # Clean up temp files
        for img_path in chart_images:
            try:
                if os.path.exists(img_path):
                    os.remove(img_path)
            except:
                pass
        
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating PDF: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None
    
# DEBUG FUNCTION - TEMPORAL PARA VERIFICAR COLUMNAS
def debug_column_names(processed_data):
    """Function to debug and show exact column names"""
    if 'inversiones_no_productivas' in processed_data:
        df = processed_data['inversiones_no_productivas']
        st.write("**Columnas exactas en Inversiones No Productivas:**")
        for i, col in enumerate(df.columns):
            st.write(f"{i+1}. '{col}' (tipo: {type(col).__name__})")
        
        # Mostrar primeras filas para entender los datos
        st.write("**Primeras 3 filas:**")
        st.dataframe(df.head(3))

# MAIN APPLICATION STATE
if 'data_initialized' not in st.session_state:
    st.session_state.data_initialized = False
    st.session_state.processed_data = None
    st.session_state.analysis_results = None
    st.session_state.template_downloaded = False

# HEADER
col_logout1, col_logout2 = st.columns([10, 1])
with col_logout2:
    if st.button("Cerrar Sesión", key="logout"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()

st.markdown("""
<div class="corporate-header">
    <div style="display: flex; align-items: center; gap: 2rem; margin-bottom: 1rem;">
        <div style="
            background: rgba(255,255,255,0.2);
            border-radius: 12px;
            padding: 12px 20px;
            font-size: 1.5rem;
            font-weight: 700;
            letter-spacing: 2px;
            border: 2px solid rgba(255,255,255,0.3);
        ">
            PROALTUS
        </div>
        <div>
            <h1 class="header-title">Análisis de Portafolio</h1>
        </div>
    </div>
    <p class="header-subtitle">Radiografía Financiera Inicial - Metodología Proaltus v4.0</p>
</div>
""", unsafe_allow_html=True)

# STATUS PANEL
col1, col2, col3 = st.columns([2, 1, 1])

with col1:
    if st.session_state.data_initialized:
        st.markdown("""
        <div class="status-indicator status-success">
            <div class="status-dot"></div>
            Sistema Activo - Radiografía Financiera Procesada
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="status-indicator status-warning">
            <div class="status-dot"></div>
            Sistema Listo - Descarga la Plantilla para Comenzar
        </div>
        """, unsafe_allow_html=True)

with col2:
    current_time = datetime.now().strftime('%H:%M:%S')
    st.markdown(f"""
    <div style="text-align: center; color: var(--medium-gray); font-family: 'JetBrains Mono', monospace;">
        <div style="font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em;">Última Actualización</div>
        <div style="font-weight: 600;">{current_time}</div>
    </div>
    """, unsafe_allow_html=True)

with col3:
    template_status = "Descargada" if st.session_state.template_downloaded else "Disponible"
    st.markdown(f"""
    <div style="text-align: center; color: var(--medium-gray); font-family: 'JetBrains Mono', monospace;">
        <div style="font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em;">Estado Plantilla</div>
        <div style="font-weight: 600;">{template_status}</div>
    </div>
    """, unsafe_allow_html=True)

# TEMPLATE AND UPLOAD SECTION
if not st.session_state.data_initialized:
    st.markdown("""
    <div class="section-container">
        <h2 style="color: #1E3A8A; margin-bottom: 1rem;">Sistema de Plantilla Excel Inteligente</h2>
        <p>Descarga nuestra plantilla Excel preconfigurada siguiendo la metodología Proaltus de radiografía financiera.</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.button("Descargar Plantilla Excel", type="primary", key="download_template"):
            try:
                with st.spinner("Preparando plantilla Excel..."):
                    try:
                        with open("Final_Plantilla_proaltus.xlsx", "rb") as template_file:
                            template_data = template_file.read()
                        
                        st.session_state.template_downloaded = True
                        
                        st.download_button(
                            label="Descargar Plantilla_Proaltus_Portafolio.xlsx",
                            data=template_data,
                            file_name=f"Final_Plantilla_proaltus_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="template_download_button"
                        )
                        
                        st.success("Plantilla lista para descarga!")
                        
                    except FileNotFoundError:
                        st.warning("Archivo de plantilla no encontrado.")
                        
            except Exception as e:
                st.error(f"Error: {str(e)}")

# UPLOAD SECTION
st.markdown("""
<div class="section-container">
    <h2 style="color: #1E3A8A; margin-bottom: 1rem;">Centro de Procesamiento de Datos</h2>
</div>
""", unsafe_allow_html=True)

if not st.session_state.data_initialized:
    uploaded_file = st.file_uploader(
        "Subir Plantilla Excel Completada",
        type=['xlsx', 'xls'],
        help="Sube tu plantilla Excel completada para realizar la radiografía financiera",
        key="main_file_uploader"
    )
    
    if uploaded_file:
        st.markdown(f"""
        <div style="background: #DBEAFE; padding: 1rem; border-radius: 8px; margin: 1rem 0;">
            <strong>Archivo Listo:</strong> {uploaded_file.name} ({uploaded_file.size / 1024 / 1024:.2f} MB)
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("Procesar Radiografía Financiera", type="primary", key="process_button"):
            with st.spinner("Procesando radiografía financiera según metodología Proaltus..."):
                try:
                    processed_data = process_uploaded_template(uploaded_file)
                    
                    if processed_data:
                        analysis_results = calculate_patrimony_kpis(processed_data)
                        
                        if analysis_results:
                            st.session_state.processed_data = processed_data
                            st.session_state.analysis_results = analysis_results
                            st.session_state.data_initialized = True
                            
                            st.success("Radiografía financiera procesada exitosamente! Cargando dashboard...")
                            st.rerun()
                        else:
                            st.error("Error calculando métricas financieras")
                    else:
                        st.error("Error procesando archivo Excel. Asegúrate de usar la plantilla proporcionada.")
                        
                except Exception as e:
                    st.error(f"Error de procesamiento: {str(e)}")
                    
else:
    st.markdown("""
    <div style="background: #059669; color: white; padding: 1rem; border-radius: 8px; text-align: center;">
        <strong>Radiografía Financiera Completada</strong> - Dashboard activo según metodología Proaltus
    </div>
    """, unsafe_allow_html=True)

# MAIN DASHBOARD
if st.session_state.data_initialized and st.session_state.analysis_results:
    
    # KPI CARDS - FOLLOWING MANUAL FORMULAS
    kpis = st.session_state.analysis_results
    
    st.markdown("""
    <div class="section-container">
        <h2 style="color: #1E3A8A; margin-bottom: 2rem;">Diagnóstico Patrimonial</h2>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_patrimony = safe_float(kpis.get('total_patrimony', 0))
        asset_count = int(kpis.get('asset_count', 0))
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Patrimonio Total</div>
            <div class="kpi-value">${total_patrimony:,.0f}</div>
            <div class="kpi-meta">{asset_count} Activos Totales</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        net_flow = safe_float(kpis.get('net_flow', 0))
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Flujo Efectivo Neto (FCN)</div>
            <div class="kpi-value">${net_flow:,.0f}</div>
            <div class="kpi-meta">Balance Mensual</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        total_income = safe_float(kpis.get('total_income', 0))
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Ingresos Totales</div>
            <div class="kpi-value">${total_income:,.0f}</div>
            <div class="kpi-meta">Base de Cálculo</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        savings_rate = safe_float(kpis.get('savings_rate', 0))
        
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Tasa de Ahorro (TA)</div>
            <div class="kpi-value">{savings_rate:.1f}%</div>
            <div class="kpi-meta">TA = (FCN / Ingresos) × 100</div>
        </div>
        """, unsafe_allow_html=True)
            
    
    # CASH FLOW ANALYSIS
    st.markdown("""
    <div class="section-container">
        <h2 style="color: #1E3A8A; margin-bottom: 2rem;">Análisis de Flujo de Efectivo Requerido</h2>
        <p style="color: #6B7280; font-size: 0.875rem;">
            Metodología Proaltus de Priorización de Gastos
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    flow_analysis = generate_cash_flow_analysis(st.session_state.processed_data)
    
    if flow_analysis:
        display_cash_flow_table(flow_analysis)
        
        st.markdown("---")
        
        # CASH FLOW GRAPHIC - Gráfica 1 from Manual
        st.markdown("""
        <div style="background: #1E3A8A; color: white; padding: 1rem; text-align: center; border-radius: 8px; margin: 2rem 0 1rem 0;">
            <h2 style="margin: 0; font-size: 1.5rem; font-weight: bold; font-family: Inter;">Gráfica 1: Flujo de Efectivo</h2>
            <p style="margin: 0.5rem 0 0 0; font-size: 0.875rem; opacity: 0.9;">Estructura Jerárquica de Ingresos</p>
        </div>
        """, unsafe_allow_html=True)
        
        create_cash_flow_graphic(flow_analysis)
        
        # EXPENSES MEKKO CHART - Gráfica 2 from Manual  
        st.markdown("### Gráfica 2: Estructura de Gastos")
        create_expenses_mekko_chart(st.session_state.processed_data)
        
        st.markdown("---")
        
        # MAINTENANCE AND TAXES CHARTS - Gráficas 3 y 4 from Manual
        col1, col2 = st.columns(2)
        
        with col1:
            create_maintenance_costs_graphic(st.session_state.processed_data)
        
        with col2:
            create_taxes_graphic(st.session_state.processed_data)
        
        st.markdown("---")
        
        # INVESTMENT CHARTS - Gráficas 6-8 from Manual
        st.markdown("""
        <div style="background: #1E3A8A; color: white; padding: 1rem; text-align: center; border-radius: 8px; margin: 2rem 0 1rem 0;">
            <h2 style="margin: 0; font-size: 1.5rem; font-weight: bold; font-family: Inter;">Gráficas 6-8: Distribución del Patrimonio</h2>
            <p style="margin: 0.5rem 0 0 0; font-size: 0.875rem; opacity: 0.9;">Análisis por Categorías de Activos</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### Valor inversiones Productivas")
            
            if 'inversiones_productivas' in st.session_state.processed_data:
                df_prod = st.session_state.processed_data['inversiones_productivas']
                
                name_col = find_exact_column(df_prod, ['Nombre del Activo'])
                valor_col = find_exact_column(df_prod, ['Valor (COP)'])
                
                if name_col and valor_col and not df_prod.empty:
                    df_valid = df_prod.copy()
                    df_valid[valor_col] = pd.to_numeric(df_valid[valor_col], errors='coerce').fillna(0)
                    df_valid = df_valid[df_valid[valor_col] > 0]
                    
                    if not df_valid.empty:
                        nombres = df_valid[name_col].tolist()
                        valores = df_valid[valor_col].tolist()
                        
                        colors_prod = ['#1E3A8A', '#3B82F6', '#60A5FA', '#93C5FD', '#DBEAFE', '#F3F4F6'][:len(nombres)]
                        
                        fig_prod = px.pie(
                            values=valores,
                            names=nombres,
                            color_discrete_sequence=colors_prod,
                            hole=0.0
                        )
                        
                        fig_prod.update_layout(
                            height=300,
                            paper_bgcolor='white',
                            font=dict(family="Inter", size=9),
                            margin=dict(l=10, r=10, t=10, b=10),
                            showlegend=True,
                            legend=dict(
                                orientation="v", 
                                x=1.02, 
                                y=0.5,
                                font=dict(size=8, family="Inter")
                            )
                        )
                        
                        fig_prod.update_traces(
                            textposition='inside',
                            textinfo='percent',
                            textfont_size=10,
                            textfont_family="Inter"
                        )
                        
                        st.plotly_chart(fig_prod, use_container_width=True)
                    else:
                        st.warning("No valid productive investment data")
                else:
                    st.warning("Required columns not found for productive investments")
            
            st.markdown("#### Inversiones Financieras por Asset Class")
            create_financial_investments_chart(st.session_state.processed_data)
            
            st.markdown("#### Inversiones Financieras por Sub Asset Class")
            create_financial_sub_asset_chart(st.session_state.processed_data)
        
        with col2:
            st.markdown("#### Valor Inversiones No Productivas") 
            
            if 'inversiones_no_productivas' in st.session_state.processed_data:
                df_no_prod = st.session_state.processed_data['inversiones_no_productivas']
                
                name_col = find_exact_column(df_no_prod, ['Nombre del Activo'])
                valor_col = find_exact_column(df_no_prod, ['Valor (COP)'])
                
                if name_col and valor_col and not df_no_prod.empty:
                    df_valid = df_no_prod.copy()
                    df_valid[valor_col] = pd.to_numeric(df_valid[valor_col], errors='coerce').fillna(0)
                    df_valid = df_valid[df_valid[valor_col] > 0]
                    
                    if not df_valid.empty:
                        nombres_np = df_valid[name_col].tolist()
                        valores_np = df_valid[valor_col].tolist()
                        
                        colors_np = ['#1E3A8A', '#60A5FA', '#10B981', '#34D399', '#6B7280', '#9CA3AF', '#D1D5DB'][:len(nombres_np)]
                        
                        fig_np = px.pie(
                            values=valores_np,
                            names=nombres_np,
                            color_discrete_sequence=colors_np,
                            hole=0.0
                        )
                        
                        fig_np.update_layout(
                            height=300,
                            paper_bgcolor='white',
                            font=dict(family="Inter", size=8),
                            margin=dict(l=10, r=10, t=10, b=10),
                            showlegend=True,
                            legend=dict(
                                orientation="v", 
                                x=1.02, 
                                y=0.5,
                                font=dict(size=7, family="Inter")
                            )
                        )
                        
                        fig_np.update_traces(
                            textposition='inside',
                            textinfo='percent',
                            textfont_size=9,
                            textfont_family="Inter"
                        )
                        
                        st.plotly_chart(fig_np, use_container_width=True)
                    else:
                        st.warning("No valid non-productive investment data")
                else:
                    st.warning("Required columns not found for non-productive investments")
            
            st.markdown("#### Gráfica 9: Distribución del Patrimonio")
            create_patrimony_mekko_chart(kpis)
        
        st.markdown("---")
        
        # CURRENCY DISTRIBUTION CHARTS
        st.markdown("""
        <div style="background: #1E3A8A; color: white; padding: 1rem; text-align: center; border-radius: 8px; margin: 2rem 0 1rem 0;">
            <h2 style="margin: 0; font-size: 1.5rem; font-weight: bold; font-family: Inter;">Distribución por Moneda</h2>
            <p style="margin: 0.5rem 0 0 0; font-size: 0.875rem; opacity: 0.9;">Análisis de Diversificación por Tipo de Moneda</p>
        </div>
        """, unsafe_allow_html=True)

        # STACKED BAR: By Asset Type and Currency
        st.markdown("### Distribución por Tipo de Activo y Moneda")
        create_currency_chart(st.session_state.processed_data)

        # NEW PIE CHART: Total by Currency Only
        st.markdown("### Valor Total por Moneda")
        create_currency_pie_chart(st.session_state.processed_data)

        st.markdown("---")
                
        # PROFITABILITY BREAKDOWN CHART
        st.markdown("### Desglose de Rentabilidad - Inversiones Financieras")
        create_profitability_breakdown_chart(st.session_state.processed_data)
        
        # EXPECTED RETURNS CHART - Gráfica 10 from Manual
        st.markdown("### Gráfica 10: Rendimiento Esperado")
        create_return_graphic()

        st.markdown("---")
        
        # GEOGRAPHIC DISTRIBUTION MAP
        st.markdown("""
        <div style="background: #1E3A8A; color: white; padding: 1rem; text-align: center; border-radius: 8px; margin: 2rem 0 1rem 0;">
            <h2 style="margin: 0; font-size: 1.5rem; font-weight: bold; font-family: Inter;">Mapa de Distribución Geográfica</h2>
            <p style="margin: 0.5rem 0 0 0; font-size: 0.875rem; opacity: 0.9;">Ubicación y Concentración de Activos</p>
        </div>
        """, unsafe_allow_html=True)
        
        create_geographic_distribution_map(st.session_state.processed_data)

        st.markdown("---")
        
        # COST COMPARISON CHART
        st.markdown("""
        <div style="background: #1E3A8A; color: white; padding: 1rem; text-align: center; border-radius: 8px; margin: 2rem 0 1rem 0;">
            <h2 style="margin: 0; font-size: 1.5rem; font-weight: bold; font-family: Inter;">Análisis de Valor Proaltus</h2>
            <p style="margin: 0.5rem 0 0 0; font-size: 0.875rem; opacity: 0.9;">Comparación de Costos de Gestión</p>
        </div>
        """, unsafe_allow_html=True)
        
        create_cost_comparison_chart(st.session_state.processed_data)

    else:
        st.error("No se pudo generar el análisis de flujo de efectivo. Verifica que los datos estén completos.")
    
    # ADDITIONAL ANALYSIS INDICATORS
    st.markdown("""
    <div class="section-container">
        <h2 style="color: #1E3A8A; margin-bottom: 2rem;">Indicadores Adicionales de Salud Financiera</h2>
    </div>
    """, unsafe_allow_html=True)
    
    if flow_analysis:
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            # Independence indicator (Equation 18)
            total_income = flow_analysis['ingresos']['total']
            passive_income = flow_analysis['ingresos']['ingresos_pasivos']
            independence_rate = (passive_income / total_income * 100) if total_income > 0 else 0
            
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-title">Independencia Financiera (IF)</div>
                <div class="kpi-value">{independence_rate:.1f}%</div>
                <div class="kpi-meta">Ingresos Pasivos / Total</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            # Maintenance burden (Equation 19)
            maintenance_cost = flow_analysis['gastos_p1']['relacionado_inversiones']
            maintenance_burden = (maintenance_cost / total_income * 100) if total_income > 0 else 0
            
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-title">Carga de Mantenimiento (CM)</div>
                <div class="kpi-value">{maintenance_burden:.1f}%</div>
                <div class="kpi-meta">Costos Mant. / Ingresos</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            # Priority 1 expenses ratio
            gp1_rate = flow_analysis['resumen']['porcentajes']['gastos_p1']
            
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-title">Gastos Prioridad 1 (GP1)</div>
                <div class="kpi-value">{gp1_rate:.1f}%</div>
                <div class="kpi-meta">Gastos Esenciales</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            # Priority 2 expenses ratio
            gp2_rate = flow_analysis['resumen']['porcentajes']['gastos_p2']
            
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-title">Gastos Prioridad 2 (GP2)</div>
                <div class="kpi-value">{gp2_rate:.1f}%</div>
                <div class="kpi-meta">Gastos Discrecionales</div>
            </div>
            """, unsafe_allow_html=True)
    
    # REPORTS AND ACTIONS
    st.markdown("""
    <div class="section-container">
        <h2 style="color: #1E3A8A; margin-bottom: 2rem;">Reportes y Exportación</h2>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if flow_analysis:
            csv_data = pd.DataFrame({
                'Concepto': ['Ingresos', 'Gastos Prioridad 1', 'Gastos Prioridad 2', 'Inversiones', 'Impuestos', 'Total Egresos', 'FCN'],
                'Monto_COP': [
                    safe_float(flow_analysis['ingresos']['total']),
                    safe_float(flow_analysis['gastos_p1']['total']),
                    safe_float(flow_analysis['gastos_p2']['total']),
                    safe_float(flow_analysis['inversiones']['total']),
                    safe_float(flow_analysis['impuestos']['total']),
                    safe_float(flow_analysis['resumen']['total_egresos']),
                    safe_float(flow_analysis['resumen']['resultado_neto'])
                ],
                'Porcentaje': [
                    100.0,
                    safe_float(flow_analysis['resumen']['porcentajes']['gastos_p1']),
                    safe_float(flow_analysis['resumen']['porcentajes']['gastos_p2']),
                    safe_float(flow_analysis['resumen']['porcentajes']['inversiones']),
                    safe_float(flow_analysis['resumen']['porcentajes']['impuestos']),
                    0.0,
                    safe_float(flow_analysis['resumen']['porcentajes']['resultado_neto'])
                ]
            }).to_csv(index=False)
            
            st.download_button(
                label="Exportar Análisis FCN",
                data=csv_data,
                file_name=f"analisis_flujo_efectivo_proaltus_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )
    
    with col2:
        if st.button("Generar Diagnóstico Completo", key="full_report"):
            st.info("Generando diagnóstico patrimonial completo según metodología Proaltus...")
    
    with col3:
        if st.button("Descargar Radiografía PDF", key="pdf_download"):
            if PDF_AVAILABLE:
                with st.spinner("Generando radiografía financiera en PDF..."):
                    pdf_data = generate_pdf_report(flow_analysis, kpis, st.session_state.processed_data)
                    if pdf_data:
                        st.download_button(
                            label="Descargar Radiografía PDF",
                            data=pdf_data,
                            file_name=f"radiografia_financiera_proaltus_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                            mime="application/pdf",
                            key="pdf_download_button"
                        )
                        st.success("Radiografía PDF generada exitosamente!")
                    else:
                        st.error("Error al generar la radiografía PDF")
            else:
                st.error("Funcionalidad PDF no disponible. Se requiere instalar reportlab.")
    
    with col4:
        if st.button("Reiniciar Sistema", key="reset_system"):
            authenticated = st.session_state.get('authenticated', False)
            for key in list(st.session_state.keys()):
                if key != 'authenticated' and key != 'page_config_set':
                    del st.session_state[key]
            st.session_state.authenticated = authenticated
            st.rerun()

    # CLIENT PROFILE SECTION ACCORDING TO MANUAL
    st.markdown("""
    <div class="section-container">
        <h2 style="color: #1E3A8A; margin-bottom: 2rem;">Perfil del Cliente y Recomendaciones</h2>
    </div>
    """, unsafe_allow_html=True)
    
    if flow_analysis:
        savings_rate = flow_analysis['resumen']['porcentajes']['resultado_neto']
        
        # Client profile classification according to manual Section 8.1
        if savings_rate > 20:
            profile = "Cliente Acumulador (Alta Tasa de Ahorro)"
            profile_desc = "Características: TA > 20%, bajo nivel de gastos discrecionales"
            recommendations = "• Diversificación de inversiones\n• Productos de mayor rentabilidad\n• Optimización fiscal avanzada"
            color = "#059669"
        elif 10 <= savings_rate <= 20:
            profile = "Cliente Equilibrado (Tasa de Ahorro Moderada)"
            profile_desc = "Características: 10% < TA < 20%, estructura de gastos balanceada"
            recommendations = "• Optimización fiscal\n• Mejora en eficiencia de inversiones\n• Balanceo de portafolio"
            color = "#F59E0B"
        elif 0 <= savings_rate < 10:
            profile = "Cliente en Optimización (Baja Tasa de Ahorro)"
            profile_desc = "Características: TA < 10%, alta carga de gastos discrecionales"
            recommendations = "• Reestructuración de gastos\n• Liquidación de activos improductivos\n• Plan de ahorro forzoso"
            color = "#D97706"
        else:
            profile = "Cliente en Situación Crítica"
            profile_desc = "Características: TA < 0%, gastos superan ingresos"
            recommendations = "• Reestructuración urgente de gastos\n• Generación de ingresos adicionales\n• Liquidación de activos no esenciales"
            color = "#DC2626"
        
        st.markdown(f"""
        <div style="background: {color}; color: white; padding: 2rem; border-radius: 12px; margin-bottom: 1rem;">
            <h3 style="margin: 0 0 1rem 0; font-size: 1.5rem; font-weight: 700;">{profile}</h3>
            <p style="margin: 0 0 1rem 0; font-size: 1rem; opacity: 0.9;">{profile_desc}</p>
            <div style="background: rgba(255,255,255,0.1); padding: 1rem; border-radius: 8px;">
                <h4 style="margin: 0 0 0.5rem 0; font-size: 1.1rem;">Recomendaciones:</h4>
                <p style="margin: 0; white-space: pre-line;">{recommendations}</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # DEBUG SECTION
    with st.expander("🔍 Debug Info - Validación de Datos", expanded=False):
        if st.session_state.processed_data:
            st.write("**Hojas procesadas según metodología:**")
            for sheet_name, df in st.session_state.processed_data.items():
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write(f"**{sheet_name}**")
                with col2:
                    st.write(f"Filas: {len(df)}")
                with col3:
                    st.write(f"Columnas: {len(df.columns)}")
                
                if st.button(f"Ver columnas de {sheet_name}", key=f"debug_{sheet_name}"):
                    st.write("Columnas disponibles:")
                    for i, col in enumerate(df.columns):
                        st.write(f"{i+1}. {col}")
        
        # DEBUG TEMPORAL PARA COLUMNAS
        if st.button("Debug Column Names", key="debug_columns"):
            debug_column_names(st.session_state.processed_data)
        
        if flow_analysis:
            st.write("**Fórmulas aplicadas según manual técnico:**")
            st.json({
                "Ecuación 1 - Patrimonio Total": f"${kpis.get('total_patrimony', 0):,.0f}",
                "Ecuación 2 - FCN": f"${flow_analysis['resumen']['resultado_neto']:,.0f}",
                "Ecuación 3 - Ingresos Totales": f"${flow_analysis['ingresos']['total']:,.0f}",
                "Ecuación 4 - Egresos Totales": f"${flow_analysis['resumen']['total_egresos']:,.0f}",
                "Ecuación 5 - Tasa de Ahorro": f"{flow_analysis['resumen']['porcentajes']['resultado_neto']:.1f}%",
                "Ecuación 6 - Conteo Activos": int(kpis.get('asset_count', 0))
            })

# VALIDATION AND WARNINGS
if st.session_state.data_initialized:
    if st.session_state.processed_data:
        warnings = []
        
        # Check patrimony
        total_patrimony = safe_float(kpis.get('total_patrimony', 0))
        if total_patrimony == 0:
            warnings.append("⚠️ Patrimonio total es $0 - Verifica los valores en COP")
        
        # Check cash flow
        if flow_analysis:
            total_income = flow_analysis['ingresos']['total']
            if total_income == 0:
                warnings.append("⚠️ No se detectaron ingresos - Verifica la hoja 'Datos adicionales'")
            
            savings_rate = flow_analysis['resumen']['porcentajes']['resultado_neto']
            if savings_rate < 0:
                warnings.append("🔴 FCN negativo - Los gastos superan los ingresos (Situación Crítica)")
            elif savings_rate < 10:
                warnings.append("🟡 Baja tasa de ahorro - Cliente en optimización requerida")
        
        # Check data completeness
        expected_sheets = ['empresas', 'inversiones_no_productivas', 'inversiones_productivas', 'inversiones_financieras', 'datos_adicionales']
        missing_sheets = [sheet for sheet in expected_sheets if sheet not in st.session_state.processed_data]
        if missing_sheets:
            warnings.append(f"⚠️ Hojas faltantes: {', '.join(missing_sheets)}")
        
        if warnings:
            st.markdown("### 🚨 Alertas del Sistema de Diagnóstico")
            for warning in warnings:
                st.warning(warning)

# FOOTER
st.markdown(f"""
<div style="margin-top: 4rem; padding: 2rem 0; text-align: center; color: #6B7280; border-top: 1px solid #E5E7EB;">
    <div style="font-size: 0.875rem; font-weight: 500; margin-bottom: 0.5rem;">
        Proaltus Dashboard de Análisis de Portafolio v4.0 - Metodología Completa
    </div>
    <div style="font-size: 0.75rem; opacity: 0.8;">
        ✅ Fórmulas según Manual Técnico • ✅ Radiografía Financiera Completa • ✅ Diagnóstico Patrimonial • ✅ Clasificación de Perfiles
    </div>
    <div style="font-size: 0.75rem; margin-top: 0.5rem; opacity: 0.6;">
        Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | 
        Sistema de gestión patrimonial profesional según metodología Proaltus
    </div>
</div>
""", unsafe_allow_html=True)

# FOOTER
st.markdown(f"""
<div style="margin-top: 4rem; padding: 2rem 0; text-align: center; color: #6B7280; border-top: 1px solid #E5E7EB;">
    <div style="font-size: 0.875rem; font-weight: 500; margin-bottom: 0.5rem;">
        Proaltus Dashboard de Análisis de Portafolio v4.0 - Metodología Completa
    </div>
    <div style="font-size: 0.75rem; opacity: 0.8;">
        ✅ Fórmulas según Manual Técnico • ✅ Radiografía Financiera Completa • ✅ Diagnóstico Patrimonial • ✅ Clasificación de Perfiles
    </div>
    <div style="font-size: 0.75rem; margin-top: 0.5rem; opacity: 0.6;">
        Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | 
        Sistema de gestión patrimonial profesional según metodología Proaltus
    </div>
</div>
""", unsafe_allow_html=True)
# 🔍 DEBUGGING DETALLADO - VALORES CALCULADOS
if st.session_state.data_initialized and flow_analysis:
    with st.expander("🔍 DEBUG: Ver Cálculos Detallados de FCN", expanded=True):
        st.markdown("### 💰 INGRESOS")
        st.json({
            "Ingreso Salarial": f"${flow_analysis['ingresos']['ingreso_salarial']:,.0f}",
            "Ingresos Pasivos": f"${flow_analysis['ingresos']['ingresos_pasivos']:,.0f}",
            "TOTAL INGRESOS": f"${flow_analysis['ingresos']['total']:,.0f}"
        })
        
        st.markdown("### 💸 EGRESOS - PRIORIDAD 1 (GP1)")
        st.json({
            "Gastos Esenciales": f"${flow_analysis['gastos_p1']['gastos_esenciales']:,.0f}",
            "Gastos Operativos": f"${flow_analysis['gastos_p1']['gastos_operativos']:,.0f}",
            "Mantenimiento Inversiones": f"${flow_analysis['gastos_p1']['relacionado_inversiones']:,.0f}",
            "TOTAL GP1": f"${flow_analysis['gastos_p1']['total']:,.0f}"
        })
        
        st.markdown("### 💸 EGRESOS - PRIORIDAD 2 (GP2)")
        st.json({
            "Gastos Varios": f"${flow_analysis['gastos_p2']['gastos_varios']:,.0f}",
            "Viajes": f"${flow_analysis['gastos_p2']['viajes']:,.0f}",
            "Lujo": f"${flow_analysis['gastos_p2']['lujo']:,.0f}",
            "TOTAL GP2": f"${flow_analysis['gastos_p2']['total']:,.0f}"
        })
        
        st.markdown("### 💸 INVERSIONES (INV)")
        st.json({
            "Pensión Voluntaria": f"${flow_analysis['inversiones']['pension_voluntaria']:,.0f}",
            "Proyectos Inmobiliarios": f"${flow_analysis['inversiones']['proyecto_inmobiliarios']:,.0f}",
            "TOTAL INV": f"${flow_analysis['inversiones']['total']:,.0f}"
        })
        
        st.markdown("### 💸 IMPUESTOS (IMP)")
        st.json({
            "Impuestos Inversiones (mensual)": f"${flow_analysis['impuestos']['impuestos_inversiones']:,.0f}",
            "Provisión Impuestos": f"${flow_analysis['impuestos']['provision_impuestos']:,.0f}",
            "TOTAL IMP": f"${flow_analysis['impuestos']['total']:,.0f}"
        })
        
        st.markdown("### 📊 RESUMEN FINAL")
        st.json({
            "TOTAL INGRESOS": f"${flow_analysis['ingresos']['total']:,.0f}",
            "TOTAL EGRESOS": f"${flow_analysis['resumen']['total_egresos']:,.0f}",
            "FCN (Ingresos - Egresos)": f"${flow_analysis['resumen']['resultado_neto']:,.0f}",
            "Tasa de Ahorro": f"{flow_analysis['resumen']['porcentajes']['resultado_neto']:.2f}%"
        })
        
        # Verificar si Costo mantenimiento se está leyendo
        if 'inversiones_no_productivas' in st.session_state.processed_data:
            df_np = st.session_state.processed_data['inversiones_no_productivas']
            costo_col = find_exact_column(df_np, ['Costo mantenimiento'])
            st.markdown("### 🔧 VERIFICACIÓN: Costos de Mantenimiento")
            if costo_col:
                st.success(f"✅ Columna encontrada: '{costo_col}'")
                total_costo_mant = df_np[costo_col].sum()
                st.write(f"**Suma de costos de mantenimiento:** ${total_costo_mant:,.0f}")
                st.dataframe(df_np[['Nombre del Activo', costo_col]].head(10))
            else:
                st.error("❌ No se encuentra la columna 'Costo mantenimiento'")

